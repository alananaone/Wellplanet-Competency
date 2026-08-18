import json
import os
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

DATA_PATH = os.path.join(os.path.dirname(__file__), "..", "data", "evaluation_data.json")
if not os.path.exists(DATA_PATH):
    DATA_PATH = os.path.join(os.path.dirname(__file__), "..", "evaluation_data.json")

with open(DATA_PATH, "r", encoding="utf-8") as f:
    all_entries = json.load(f)

OUTPUT_DIR = os.path.join(os.path.dirname(__file__), "..", "reports_excel")
os.makedirs(OUTPUT_DIR, exist_ok=True)

# Exact colors from VVN reference
COLOR_TITLE_BG = "F4CCCC"       # Pink blush
COLOR_SECTION_BG = "FCE5CD"     # Peach cream
COLOR_WHITE = "FFFFFF"
COLOR_FEEDBACK_BG = "FFF2D6"    # Warm highlight for feedback/supervisor ratings
COLOR_COLLAB_BG = "CFE2F3"      # Soft pastel blue for 協作狀況
COLOR_CULTURE_BG = "D9D2E9"     # Soft pastel purple for 文化實踐
COLOR_AVG_BG = "E4ECD3"         # Soft green for 平均得分
COLOR_RATING_TITLE_BG = "FBE4EA"# Pink header for rating table
COLOR_RATING_HDR_BG = "E8638A"  # Deep rose header for rating table
COLOR_L5_BG = "E4EDF7"          # Amazing! soft sky blue
COLOR_L4_BG = "E3F1E6"          # Good soft mint green
COLOR_L3_BG = "FFF7E0"          # Keep soft warm yellow (fixed!)
COLOR_L2_BG = "FCE8EC"          # Grow soft blush pink
COLOR_L1_BG = "FCE8EC"          # Start soft blush pink

thin_side = Side(border_style="thin", color="D7CCC8")
thin_border = Border(top=thin_side, left=thin_side, right=thin_side, bottom=thin_side)

font_main_title = Font(name="微軟正黑體", size=13, bold=True, color="3E2723")
font_sub_header = Font(name="微軟正黑體", size=10.5, bold=True, color="4E342E")
font_table_hdr = Font(name="微軟正黑體", size=10, bold=True, color="4E342E")
font_body = Font(name="微軟正黑體", size=9.5, color="2D2323")
font_body_bold = Font(name="微軟正黑體", size=9.5, bold=True, color="2D2323")
font_rating_hdr = Font(name="微軟正黑體", size=10, bold=True, color="FFFFFF")

align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
align_left = Alignment(horizontal="left", vertical="top", wrap_text=True)
align_header = Alignment(horizontal="center", vertical="center", wrap_text=True)

# Official Org Structure & Titles
JOB_ROLES_MAP = {
    "張希慈": "執行長",
    "陳泳璇": "行政經理",
    "林文琇": "美感設計師",
    "胡喻翔": "專案經理",
    "張芳媐": "營運經理兼執行長特助",
    "何維安": "品牌經理",
    "姚品瑄": "部門儲備主管",
    "薛筑瑄": "專案經理",
    "戴佑珍": "專案經理"
}

SUPERVISOR_TEAMS = {
    "何維安": ["林文琇"],
    "姚品瑄": ["薛筑瑄", "戴佑珍"],
    "張希慈": ["何維安", "陳泳璇", "張芳媐", "姚品瑄", "胡喻翔"],
    "張希慈_執行長": ["張希慈"]
}

MEMBER_SUPERVISOR_MAP = {
    "何維安": "張希慈", "陳泳璇": "張希慈", "張芳媐": "張希慈", "姚品瑄": "張希慈", "胡喻翔": "張希慈",
    "林文琇": "何維安", "薛筑瑄": "姚品瑄", "戴佑珍": "姚品瑄", "張希慈": "董事會"
}

ALL_MEMBERS = ["何維安", "陳泳璇", "張芳媐", "姚品瑄", "胡喻翔", "林文琇", "薛筑瑄", "戴佑珍", "張希慈"]

# Comprehensive Competency Definitions extracted from 題庫總表
COMPETENCY_DEFINITIONS = {
    # 執行長
    "策略決策與組織方向設定": "1. 能將願景使命轉譯為可執行的年度策略目標，讓各部門知道自己的工作與組織方向的關係。\n2. 面對多個可行方向時，能說明選擇與捨棄的判準，而非僅宣布結論。\n3. 能整合各部門資料與不同利害關係人需求，提出資源分配邏輯。\n4. 能在環境變動時重新排序優先順序，並向團隊說明調整的理由。",
    "組織治理與財務風險管理": "1. 能維持董事會、主管機關與內部決策機制的正常運作與資訊透明。\n2. 能在重大決策前辨識法規、財務或聲譽風險，並提出因應方式。\n3. 能掌握組織整體財務狀況，在收支結構出現警訊時提前處理。\n4. 能建立可被檢驗的決策紀錄與授權範圍。",
    "關鍵利害關係人關係建立與維繫": "1. 能主動開發新的資源來源（人力、廠商、捐贈、補助、媒體）。\n2. 能維繫既有重要關係人的長期互動，而非僅在需要時才聯繫。\n3. 能在合作條件協商中兼顧組織利益與夥伴關係。\n4. 能為組織節省成本或爭取到原本不會有的資源。",
    "公關、演講與媒體關係": "1. 對外發言能清楚傳達基金會的理念，且與內部說法一致。\n2. 能因應不同受眾（媒體、捐款人、學校、政府）調整表達方式而不失核心訊息。\n3. 能在爭議或負面訊息出現時妥適回應，降低對組織的傷害。\n4. 公開露出能實際轉化為認識、信任或資源。",
    "核心團隊培養與組織文化建立": "1. 能辨識主管層的發展需求，給予具體的授權與練習機會。\n2. 能在關鍵決策上讓主管參與，而非全部由自己決定。\n3. 能在團隊出現文化落差時，直接處理而非迴避。\n4. 能建立讓夥伴敢於表達不同意見的討論場域。",

    # 行政經理
    "人事薪資與人資系統管理": "1. 準確執行薪資計算、勞健保與退休金投保作業。\n2. 維護人事資料與差勤紀錄的完整性與即時性。\n3. 確保人資系統的資料一致、可查、可交接。\n4. 能主動掌握法令變更並即時更新作業流程。",
    "法規與政府公文管理": "1. 掌握勞動法規、主管機關函文與法規異動。\n2. 於期限內完成公文收發、申報、陳報與回覆。\n3. 將法規與主管機關要求轉化為組織內部可執行的作業方式。\n4. 維護公文檔案與法定紀錄之完整保存。",
    "董事會與治理作業執行": "1. 依主管機關與組織章程要求，完成董事會會前準備、議程編排與資料寄送。\n2. 執行現場會議支援與精確完備之會議紀錄撰寫。\n3. 確實完成後續主管機關備查、變更登記等法定程序。\n4. 確保基金會法人治理文件與印鑑管理之合規與安全。",
    "財務核銷與內控執行": "1. 依核銷規範與補助專案要求審核單據憑證。\n2. 嚴格控管核銷流程、付款時程與合規性，及早辨識異常與風險。\n3. 確保帳務資料正確、透明且經得起內外部查核。\n4. 主動向專案同仁說明核銷規範與提供改善建議。",
    "總務採購與行政庶務管理": "1. 負責採購評估、廠商聯繫、詢比議價與發包作業。\n2. 建立並維持財產清冊、辦公物資與設備耗材管理。\n3. 維持辦公空間環境運作安全與日常行政庶務運作。\n4. 在成本控管、品質與時效之間做出合理且具說服力的判斷。",

    # 美感設計師
    "品牌識別系統設計與維護": "1. 設計 LOGO、標準字、品牌顏色與 CIS，確保品牌在名片、官網、招牌、包裝等載體一致。\n2. 能理解並實踐品牌定位與目標客群，將抽象策略轉化為具體視覺元素。\n3. 能建立並維護可供他人使用的視覺規範或模板。\n4. 能在既有識別系統下延伸出新的活動主題視覺，而不破壞一致性。",
    "視覺與美感設計實務": "1. 精通 Photoshop、Illustrator、Canva 等軟體，運用留白與字體設計建立層次感。\n2. 理解紙材、顏色模式（CMYK／RGB）及特殊加工，確保設計落地不失真。\n3. 能完成現場活動印刷品的設計、輸出、佈置與採購。\n4. 能在預算與時程限制下提出可行的材質與工法方案。",
    "需求釐清與創意提案": "1. 能在接案時主動問清楚使用情境、受眾與成功標準。\n2. 能蒐集參考資料並提出一個以上的方向供選擇。\n3. 能說明設計選擇背後的理由，而非只呈現成品。\n4. 能在收到修改意見時辨識真正的問題，而非逐條照改。",

    # 專案經理 (共通 & CGL / Soul LAB)
    "專案企劃與現場執行": "1. 事前與專案負責人及關係人確認任務細節，熟悉角色任務並預防突發狀況。\n2. 事中能獨立完成分內任務，並依時間與重要程度排序完成順序；能主動觀察服務對象、關係人及夥伴的需求，適時給予協助。\n3. 能將企劃內容落地成可執行的方案（流程表、物資、人力配置），具備從 0 到 1 企劃課程、活動與體驗的能力。",
    "專案時程與預算規劃管理": "1. 能盤點專案從起始到結尾的所有工作項目，評估各項所需時程。\n2. 能在專案中依實際狀況更新時程，即時與關係人同步。\n3. 能在專案啟動前合理分配各項目預算，過程中進行財務紀錄與評估。\n4. 能在時程或預算出現異常時排解問題，仍能在範圍內完成目標。",
    "需求研究與方案迭代": "1. 能設計並實施多元研究方法（深度訪談、問卷調查、現場觀察），主動蒐集服務對象的好奇與疑惑。\n2. 能識別、篩選並整理出有效資訊，不被個別意見帶著走；能從資料中歸納出使用者需求與行為模式。\n3. 能針對研究結果實際調整課程或方案，於下一次專案執行前完成迭代。",
    "外部夥伴關係經營": "1. 能向外部關係人（學校、講師、合作單位）提出清楚的合作需求與預期成果。\n2. 能考量對方的排程與習慣，在對方舒適的狀態下共同完成任務。\n3. 能讓實習生、志工與小組長有效分擔執行性工作，騰出正職做規劃型工作的時間。\n4. 能在外部夥伴結束合作時，讓他們對基金會留下正向感受。",
    "多元教學設計與現場引導": "1. 能依課程目標選擇合適的教學手段，並說明選擇理由。\n2. 能設計出有層次的課程流程（暖身、主活動、收斂、回顧）。\n3. 能在課程中安排讓孩子動手、動身體或表達的環節。\n4. 能說明自己的教育理念，並在課程設計中看得到這個理念。\n5. 能在現場主持引導課程進行，掌握節奏與時間。\n6. 能在課後回顧中辨識哪些設計有效、哪些需要調整。",
    "社群經營與培力需求回應": "1. 能創造與 Soul LAB 文化一致的社群氛圍（溫暖、療癒、支持）。\n2. 能讓服務對象在社群中彼此創造良好的社群氛圍（互助、回饋）。\n3. 主動觀察服務對象在社群內部的狀態、和 Soul LAB 互動的狀態。\n4. 能順暢和服務對象溝通、傳遞資訊，長期維持良好的互動關係。",

    # 營運經理兼執行長特助
    "組織營運流程設計與優化": "1. 能依組織目標規劃年度營運重點與執行節奏。\n2. 能設計並優化跨部門協作流程，減少重工與資訊落差。\n3. 能建立會議與資訊流通機制，讓決策所需資訊在會前就到位。\n4. 能建立並追蹤營運成效指標。\n5. 能在多專案並行時維持整體運作秩序。",
    "制度文件與 SOP 建置": "1. 能辨識哪些反覆發生的工作需要制度化，並產出可被他人使用的文件。\n2. 制度文件的用語與結構清楚，不需作者在旁說明也能執行。\n3. 能定期檢視既有制度是否還符合現況並進行更新。\n4. 能確認制度內容符合勞動法規與主管機關要求。",
    "人力資源策略與選用育留": "1. 能依組織發展階段規劃人力配置策略與職務說明書。\n2. 能設計招募與甄選流程，辨識與組織價值契合的人才。\n3. 能建立新人訓練與在職培訓機制。\n4. 能設計績效回饋與薪酬調整制度。\n5. 能評估人才流動對組織穩定與策略目標的影響。",
    "員工關係處理與勞動合規": "1. 能在衝突或申訴發生時依既定程序處理，並保護當事人的權益。\n2. 能依法辦理任用、留停、離職等人事程序，文件完備。\n3. 能主動掌握法規變動並調整內部作法。\n4. 能規劃並執行員工福利制度（聚會、旅行、年度活動、津貼）。",
    "組織文化落實與制度轉化": "1. 能把抽象的文化語言轉成具體可執行的制度或會議設計。\n2. 能設計讓夥伴實際練習文化行為的場合（工作坊、儀式、回饋機制）。\n3. 能觀察到文化落差的訊號，並提出結構性的處理方式。\n4. 能在制度設計時兼顧文化一致性，而非只求效率。",

    # 品牌經理
    "品牌定位與外部溝通一致性": "1. 能清楚界定組織使命、核心價值與品牌定位。\n2. 能建立並維護品牌識別系統（含視覺與敘事架構）。\n3. 能確保各部門對外傳遞一致的品牌語言與敘事。\n4. 能透過資訊透明與成效揭露建立公信力。",
    "行銷策略與議題倡議": "1. 能規劃整合行銷與議題倡議策略以提升品牌能見度。\n2. 能透過數位媒體與實體活動強化品牌認知與參與。\n3. 能透過數據分析與市場洞察調整品牌策略方向。\n4. 能運用品牌影響力吸引企業、政府與多元資源支持。",
    "品牌活動策劃與策展敘事": "1. 能策劃展覽、論壇或公共活動，將理念轉化為具體體驗。\n2. 能設計策展敘事脈絡，使受眾理解品牌價值。\n3. 能監督活動執行品質並管理協力廠商。\n4. 能在活動後評估成效並累積可複製的經驗。",
    "內部品牌管理與雇主品牌": "1. 能將品牌核心價值轉化為制度與行為準則。\n2. 能設計內部溝通機制以強化品牌認同。\n3. 能建立雇主品牌形象以吸引理念契合的人才。\n4. 能協助團隊成員成為品牌的倡議者與代言人。",
    "品牌危機與聲譽風險處理": "1. 能預判可能引發爭議的訊息或行動，事前調整。\n2. 能在負面訊息出現時快速判斷回應層級與方式。\n3. 能預防內外品牌形象失調並進行調整。\n4. 能建立危機處理的內部流程與發言原則。",

    # 部門儲備主管
    "部門策略規劃與專案組合管理": "1. 能制定部門年度發展策略，並確保各專案符合組織願景與策略方向。\n2. 能評估新專案的可行性與資源需求，做出啟動、調整或暫停的判斷。\n3. 能建立專案成效指標與定期檢核機制。\n4. 能在各專案之間協調資源與優先順序。",
    "專案經理管理與培育": "1. 能明確界定每位專案經理的權責範圍與目標，不留模糊地帶。\n2. 能定期進行專案檢核與回顧會議，而非只在出問題時介入。\n3. 能辨識專案經理在職能上的優勢與待加強處，給予具體回饋與指導。\n4. 能在專案經理遇到困難時給予支持，並協助排除跨部門或外部障礙。",
    "部門預算與資源配置管理": "1. 能統籌部門年度預算規劃，並在執行中控管落差。\n2. 能依專案優先順序分配人力與資源，並說明分配理由。\n3. 能監控部門整體投入與產出效益。\n4. 能在資源不足時做出取捨，而非平均分配。",
    "跨 LAB 專業掌握（CGL 教育專業與 SL 社群培力）": "1. 能理解 CGL 的教育設計邏輯（多元教學手段、現場引導、與兒童互動），並據以檢核課程品質。\n2. 能理解 Soul LAB 的社群培力邏輯（服務對象需求辨識、最小行動引導、社群氛圍營造），並據以檢核方案設計。\n3. 能辨識兩個 LAB 在方法上的差異，不用單一標準要求兩邊。\n4. 能促成兩個 LAB 之間的經驗交流與資源共用。",
    "利害關係人管理與衝突協調": "1. 能與外部合作單位及專案利害關係人進行有效溝通以促進專案運作。\n2. 能擔任團隊的溝通中心，促進成員互助合作、建立默契與信任感。\n3. 能有效化解專案團隊運作的各種危機。\n4. 能應用協商策略促進利害關係人支持專案的運作。"
}

ROLE_COMPETENCIES_MAP = {
    "執行長": ["策略決策與組織方向設定", "組織治理與財務風險管理", "關鍵利害關係人關係建立與維繫", "公關、演講與媒體關係", "核心團隊培養與組織文化建立"],
    "行政經理": ["人事薪資與人資系統管理", "法規與政府公文管理", "董事會與治理作業執行", "財務核銷與內控執行", "總務採購與行政庶務管理"],
    "美感設計師": ["品牌識別系統設計與維護", "視覺與美感設計實務", "需求釐清與創意提案"],
    "專案經理_CGL": ["專案企劃與現場執行", "專案時程與預算規劃管理", "需求研究與方案迭代", "外部夥伴關係經營", "多元教學設計與現場引導"],
    "專案經理_SL": ["專案企劃與現場執行", "專案時程與預算規劃管理", "需求研究與方案迭代", "外部夥伴關係經營", "社群經營與培力需求回應"],
    "營運經理兼執行長特助": ["組織營運流程設計與優化", "制度文件與 SOP 建置", "人力資源策略與選用育留", "員工關係處理與勞動合規", "組織文化落實與制度轉化"],
    "品牌經理": ["品牌定位與外部溝通一致性", "行銷策略與議題倡議", "品牌活動策劃與策展敘事", "內部品牌管理與雇主品牌", "品牌危機與聲譽風險處理"],
    "部門儲備主管": ["部門策略規劃與專案組合管理", "專案經理管理與培育", "部門預算與資源配置管理", "跨 LAB 專業掌握（CGL 教育專業與 SL 社群培力）", "利害關係人管理與衝突協調"]
}

MEMBER_COMPETENCY_KEY = {
    "張希慈": "執行長",
    "陳泳璇": "行政經理",
    "林文琇": "美感設計師",
    "胡喻翔": "專案經理_CGL",
    "張芳媐": "營運經理兼執行長特助",
    "何維安": "品牌經理",
    "姚品瑄": "部門儲備主管",
    "薛筑瑄": "專案經理_CGL",
    "戴佑珍": "專案經理_SL"
}

PEER_QUESTIONS = [
    ["Q24", "合作狀況", "合作狀況（溝通順暢度、資訊透明度、承諾事項達成率）", "q24_cooperation"],
    ["Q25", "工作品質", "注重細節（工作成果之品質與細緻度）", "q25_detail_oriented"],
    ["Q26", "專案進度", "準時完成（專案進度與時程掌控）", "q26_on_time"],
    ["Q27", "變動彈性", "靈活調整（面對臨時變動與突發狀況的彈性）", "q27_flexibility"],
    ["Q28", "追蹤承諾", "追蹤承諾（主動回報進度與承諾事項追蹤）", "q28_follow_up"],
    ["Q29", "透明溝通", "說明決策依據（決策與工作方法透明度）", "q29_transparency"],
    ["Q30", "多元文化", "【多元】接受不同意見（能開放聆聽反對觀點）", "q30_open_to_opposing"],
    ["Q31", "多元文化", "【多元】提出建設性觀點（在討論中給予實質建議）", "q31_constructive_opinions"],
    ["Q32", "實驗文化", "【實驗】開放調整（透過回饋持續迭代優化）", "q32_growth_mindset"],
    ["Q33", "信任文化", "【信任】分享經驗與資源（主動協助夥伴）", "q33_share_knowledge"],
    ["Q34", "可持續文化", "【可持續】讚美與肯定同儕（經常給予夥伴正向激勵）", "q34_praise_peers"],
    ["Q35", "可持續文化", "【可持續】尊重工作界線（維護彼此身心健康與邊界）", "q35_boundary_respect"],
    ["Q36", "NPS推薦", "NPS 推薦度（向他人推薦與此夥伴共事的意願）", "q36_nps_recommend"],
]

def style_cell_range(ws, start_row, start_col, end_row, end_col, font=None, fill_hex=None, alignment=None):
    fill = PatternFill(start_color=fill_hex, end_color=fill_hex, fill_type="solid") if fill_hex else None
    for r in range(start_row, end_row + 1):
        for c in range(start_col, end_col + 1):
            cell = ws.cell(row=r, column=c)
            if font: cell.font = font
            if fill: cell.fill = fill
            cell.border = thin_border
            if alignment: cell.alignment = alignment

def add_rating_standard_block(ws, start_r):
    # 三、評級標準
    r = start_r
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=4)
    ws.cell(row=r, column=1, value="三、評級標準（不對員工公布具體分數與總分，只回饋評級）")
    style_cell_range(ws, r, 1, r, 4, Font(name="微軟正黑體", size=11, bold=True, color="8C1D40"), COLOR_RATING_TITLE_BG, Alignment(horizontal="left", vertical="center", indent=1))
    ws.row_dimensions[r].height = 24
    r += 1

    headers = ["評級（最終呈現）", "Level (分數落點）", "定義", "回饋方式參考"]
    for i, h in enumerate(headers):
        ws.cell(row=r, column=i+1, value=h)
    style_cell_range(ws, r, 1, r, 4, font_rating_hdr, COLOR_RATING_HDR_BG, align_header)
    ws.row_dimensions[r].height = 24
    r += 1

    standards = [
        ("Amazing!", "L5 (9.0-10.0)", "遠超職位期待，表現為團隊之標竿與典範。", "讚賞其突出貢獻，探討經驗複製機制。", COLOR_L5_BG),
        ("Good", "L4 (7.0-8.9)", "優於職位期待，持續展現高標準成果。", "肯定並具體指出是哪些行為讓它超出標準，並設定具挑戰性的下一步目標。", COLOR_L4_BG),
        ("Keep", "L3 (5.0-6.9)", "符合職位門檻，展現穩定的工作交付。", "確認穩定度，維持既有節奏，指出下一階可以再往前的地方，選一到兩項深化", COLOR_L3_BG),
        ("Grow", "L2 (3.0-4.9)", "部分符合，部分能力/行為仍在建立階段。", "說明落差在哪，聚焦一項具體可練習的行為，納入下一期 IDP，設定可觀察的行為指標", COLOR_L2_BG),
        ("Start", "L1（1.0-2.9)", "目前的具體事證還看不到這項職能，或事證與職能要求落差明顯。", "說明落差在哪，聚焦一項具體可練習的行為，納入下一期 IDP，設定可觀察的行為指標", COLOR_L1_BG),
    ]

    for title, lvl, desc, fb, bg_hex in standards:
        ws.cell(row=r, column=1, value=title)
        ws.cell(row=r, column=2, value=lvl)
        ws.cell(row=r, column=3, value=desc)
        ws.cell(row=r, column=4, value=fb)

        style_cell_range(ws, r, 1, r, 4, font_body, bg_hex, align_left)
        ws.cell(row=r, column=1).alignment = align_center
        ws.cell(row=r, column=1).font = font_body_bold
        ws.cell(row=r, column=2).alignment = align_center
        ws.row_dimensions[r].height = 28
        r += 1

    return r

def add_subordinate_supervisor_sheet(wb, member_name):
    sheet_title = f"【{member_name}】自評vs主管評"
    ws = wb.create_sheet(title=sheet_title)
    ws.views.sheetView[0].showGridLines = True
    job_role = JOB_ROLES_MAP.get(member_name, "專案經理")
    sup_name = MEMBER_SUPERVISOR_MAP.get(member_name, "主管")

    peer_records = [e for e in all_entries if e["relation"] == "同事" and e["target"] == member_name]
    num_peers = len(peer_records)

    self_entry = next((e for e in all_entries if e["relation"] == "自評" and e["target"] == member_name), None)
    has_self = bool(self_entry and self_entry.get("self_eval"))
    se = self_entry.get("self_eval") if has_self else None

    # Title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=6)
    ws.cell(row=1, column=1, value=f"好好星球文化基金會 360 年中成長評估 - 【{member_name}】部屬自評與主管評核對照表（主管專用）")
    style_cell_range(ws, 1, 1, 1, 6, font_main_title, COLOR_TITLE_BG, Alignment(horizontal="left", vertical="center", indent=1))
    ws.row_dimensions[1].height = 30

    # Metadata
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=6)
    status_str = "已自評" if has_self else "尚未自評"
    ws.cell(row=2, column=1, value=f"評估對象：{member_name}（{job_role}） ｜ 直屬主管：{sup_name} ｜ 同儕樣本：{num_peers} 位 ｜ 自評狀態：{status_str}")
    style_cell_range(ws, 2, 1, 2, 6, font_sub_header, COLOR_SECTION_BG, Alignment(horizontal="left", vertical="center", indent=1))
    ws.row_dimensions[2].height = 24

    r = 3
    # 1. 組織文化 (一列一個面向，文字全部公布)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    ws.cell(row=r, column=1, value="一、組織文化實踐：部屬自評實例 vs 主管評核回饋（一列一個面向）")
    style_cell_range(ws, r, 1, r, 6, font_sub_header, COLOR_SECTION_BG, Alignment(horizontal="left", vertical="center", indent=1))
    ws.row_dimensions[r].height = 24
    r += 1

    cult_headers = ["文化面向", "文化定義與行為指引", "部屬自評實例 (STAR)", "部屬自評", "主管評定", "主管評核回饋與觀察"]
    for i, h in enumerate(cult_headers):
        ws.cell(row=r, column=i+1, value=h)
    style_cell_range(ws, r, 1, r, 6, font_table_hdr, COLOR_SECTION_BG, align_header)
    ws.row_dimensions[r].height = 24
    r += 1

    cult_rows = [
        ("【信任】", "獨立行動與決策、主動協作、雙向溝通，在高彈性下給予彼此信任", se.get("values", {}).get("信任", "（部屬未填寫）") if has_self else "（部屬未填寫）"),
        ("【多元】", "尊重差異、多元工作方法、主動表達不同觀點與想法", se.get("values", {}).get("多元", "（部屬未填寫）") if has_self else "（部屬未填寫）"),
        ("【實驗】", "透過開放心態嘗試修正與反思，勇於檢討及給予回饋", se.get("values", {}).get("實驗", "（部屬未填寫）") if has_self else "（部屬未填寫）"),
        ("【可持續】", "內在韌性、自我照顧、彈性的人際與工作邊界", se.get("values", {}).get("可持續", "（部屬未填寫）") if has_self else "（部屬未填寫）"),
    ]

    for c_title, c_desc, c_self in cult_rows:
        ws.cell(row=r, column=1, value=c_title)
        ws.cell(row=r, column=2, value=c_desc)
        ws.cell(row=r, column=3, value=c_self) # 全部公布文字
        ws.cell(row=r, column=4, value="")
        ws.cell(row=r, column=5, value="")
        ws.cell(row=r, column=6, value="【待主管填寫回饋】")

        style_cell_range(ws, r, 1, r, 6, font_body, COLOR_WHITE, align_left)
        ws.cell(row=r, column=1).alignment = align_center
        ws.cell(row=r, column=1).font = font_body_bold
        ws.cell(row=r, column=4).alignment = align_center
        ws.cell(row=r, column=5).alignment = align_center
        ws.cell(row=r, column=6).fill = PatternFill(start_color=COLOR_FEEDBACK_BG, end_color=COLOR_FEEDBACK_BG, fill_type="solid")
        ws.row_dimensions[r].height = max(30, min(120, len(c_self or "") // 35 * 18 + 18))
        r += 1

    r += 1
    # 2. 專業職能 (文字全部公布，等級統一寫「尚不公布」以避免影響主管評分)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=6)
    ws.cell(row=r, column=1, value=f"二、專業職能：自評 vs 主管評分並列對照【{job_role}】（逐項比對認知差異）")
    style_cell_range(ws, r, 1, r, 6, font_sub_header, COLOR_SECTION_BG, Alignment(horizontal="left", vertical="center", indent=1))
    ws.row_dimensions[r].height = 24
    r += 1

    comp_headers = ["職能項目", "職能定義與說明", "部屬自評實例 (STAR)", "部屬自評", "***主管評定 (L1~L5)", "***主管評語與回饋 (Feedback)"]
    for i, h in enumerate(comp_headers):
        ws.cell(row=r, column=i+1, value=h)
    style_cell_range(ws, r, 1, r, 6, font_table_hdr, COLOR_SECTION_BG, align_header)
    ws.row_dimensions[r].height = 24
    r += 1

    dv = DataValidation(type="list", formula1='"Amazing! (Lv5),Good (L4),Keep (L3),Grow (L2),Start (Lv1),尚未評分"', allow_blank=True)
    dv.error ='請從選單選擇等級：Amazing! (Lv5)、Good (L4)、Keep (L3)、Grow (L2)、Start (Lv1)'
    dv.errorTitle = '評分無效'
    dv.prompt = '請選取評級：Amazing! (Lv5)、Good (L4)、Keep (L3)、Grow (L2)、Start (Lv1)'
    dv.promptTitle = '等級選單'
    ws.add_data_validation(dv)

    role_key = MEMBER_COMPETENCY_KEY.get(member_name, "專案經理_CGL")
    comp_titles = ROLE_COMPETENCIES_MAP.get(role_key, [])

    for c_title in comp_titles:
        c_def = COMPETENCY_DEFINITIONS.get(c_title, "核心專業職能")
        self_ans = "（部屬未填寫自評實例）"
        if has_self and se.get("competencies"):
            item = next((x for x in se["competencies"] if x.get("title") == c_title or c_title in x.get("title", "")), None)
            if item and item.get("answer"):
                self_ans = item["answer"]

        ws.cell(row=r, column=1, value=c_title)
        ws.cell(row=r, column=2, value=c_def)
        ws.cell(row=r, column=3, value=self_ans) # 文字全部公布
        ws.cell(row=r, column=4, value="尚不公布") # 等級統一寫「尚不公布」避免影響主管評分
        ws.cell(row=r, column=5, value="尚未評分")
        ws.cell(row=r, column=6, value="【待主管填寫回饋】")

        style_cell_range(ws, r, 1, r, 6, font_body, COLOR_WHITE, align_left)
        ws.cell(row=r, column=1).font = font_body_bold
        ws.cell(row=r, column=4).alignment = align_center
        ws.cell(row=r, column=4).font = font_body_bold
        ws.cell(row=r, column=5).alignment = align_center
        ws.cell(row=r, column=5).font = font_body_bold
        ws.cell(row=r, column=5).fill = PatternFill(start_color=COLOR_FEEDBACK_BG, end_color=COLOR_FEEDBACK_BG, fill_type="solid")
        ws.cell(row=r, column=6).fill = PatternFill(start_color=COLOR_FEEDBACK_BG, end_color=COLOR_FEEDBACK_BG, fill_type="solid")

        dv.add(ws.cell(row=r, column=5))
        ws.row_dimensions[r].height = max(40, min(140, len(c_def) // 25 * 18 + 18))
        r += 1

    r += 1
    # 3. 評級標準 (三、評級標準)
    r = add_rating_standard_block(ws, r)

    ws.column_dimensions["A"].width = 16
    ws.column_dimensions["B"].width = 45
    ws.column_dimensions["C"].width = 45
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 20
    ws.column_dimensions["F"].width = 42

def add_subordinate_peer_sheet(wb, member_name):
    sheet_title = f"【{member_name}】同儕評"
    ws = wb.create_sheet(title=sheet_title)
    ws.views.sheetView[0].showGridLines = True
    job_role = JOB_ROLES_MAP.get(member_name, "專案經理")
    sup_name = MEMBER_SUPERVISOR_MAP.get(member_name, "主管")

    peer_records = [e for e in all_entries if e["relation"] == "同事" and e["target"] == member_name]
    num_peers = len(peer_records)

    self_entry = next((e for e in all_entries if e["relation"] == "自評" and e["target"] == member_name), None)
    has_self = bool(self_entry and self_entry.get("self_eval"))
    se = self_entry.get("self_eval") if has_self else None

    # Title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=7)
    ws.cell(row=1, column=1, value=f"好好星球文化基金會 360 年中成長評估 - 【{member_name}】部屬同儕評估與自評綜合報告（主管專用）")
    style_cell_range(ws, 1, 1, 1, 7, font_main_title, COLOR_TITLE_BG, Alignment(horizontal="left", vertical="center", indent=1))
    ws.row_dimensions[1].height = 30

    # Metadata
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=7)
    status_str = "已自評" if has_self else "尚未自評"
    ws.cell(row=2, column=1, value=f"評估對象：{member_name}（{job_role}） ｜ 直屬主管：{sup_name} ｜ 同儕填答樣本：{num_peers} 位 ｜ 自評狀態：{status_str}")
    style_cell_range(ws, 2, 1, 2, 7, font_sub_header, COLOR_SECTION_BG, Alignment(horizontal="left", vertical="center", indent=1))
    ws.row_dimensions[2].height = 24

    r = 3
    # 一、統計分數
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    ws.cell(row=r, column=1, value="一、統計分數：同儕量化評估分析（平均分、最高分、最低分）")
    style_cell_range(ws, r, 1, r, 7, font_sub_header, COLOR_SECTION_BG, Alignment(horizontal="left", vertical="center", indent=1))
    ws.row_dimensions[r].height = 24
    r += 1

    headers = ["題號", "評估面向", "評估項目與題目說明", "同儕平均得分", "最高分", "最低分", "給分明細"]
    for i, h in enumerate(headers):
        ws.cell(row=r, column=i+1, value=h)
    style_cell_range(ws, r, 1, r, 7, font_table_hdr, COLOR_SECTION_BG, align_header)
    ws.row_dimensions[r].height = 24
    r += 1

    collab_scores = []
    culture_scores = []

    for qNo, qCat, qDesc, qKey in PEER_QUESTIONS:
        scores = [r_entry["peer_eval"][qKey] for r_entry in peer_records if r_entry.get("peer_eval") and r_entry["peer_eval"].get(qKey) is not None]
        avg = round(sum(scores) / len(scores), 2) if scores else None
        best = max(scores) if scores else None
        worst = min(scores) if scores else None
        detail_str = ", ".join([str(s) for s in scores]) if scores else "—"

        is_collab = int(qNo.replace("Q","")) <= 29
        cat_bg = COLOR_COLLAB_BG if is_collab else COLOR_CULTURE_BG
        if qNo == "Q36": cat_bg = COLOR_WHITE

        if is_collab and avg is not None:
            collab_scores.extend(scores)
        elif not is_collab and qNo != "Q36" and avg is not None:
            culture_scores.extend(scores)

        ws.cell(row=r, column=1, value=qNo)
        ws.cell(row=r, column=2, value=qCat)
        ws.cell(row=r, column=3, value=qDesc)
        ws.cell(row=r, column=4, value=avg if avg is not None else "—")
        ws.cell(row=r, column=5, value=best if best is not None else "—")
        ws.cell(row=r, column=6, value=worst if worst is not None else "—")
        ws.cell(row=r, column=7, value=detail_str)

        style_cell_range(ws, r, 1, r, 7, font_body, COLOR_WHITE, align_left)
        ws.cell(row=r, column=1).alignment = align_center
        ws.cell(row=r, column=1).fill = PatternFill(start_color=cat_bg, end_color=cat_bg, fill_type="solid")
        ws.cell(row=r, column=2).alignment = align_center
        ws.cell(row=r, column=2).fill = PatternFill(start_color=cat_bg, end_color=cat_bg, fill_type="solid")
        ws.cell(row=r, column=4).alignment = align_center
        ws.cell(row=r, column=4).font = font_body_bold
        ws.cell(row=r, column=4).fill = PatternFill(start_color=COLOR_AVG_BG, end_color=COLOR_AVG_BG, fill_type="solid")
        ws.cell(row=r, column=5).alignment = align_center
        ws.cell(row=r, column=6).alignment = align_center
        ws.cell(row=r, column=7).alignment = align_center
        ws.row_dimensions[r].height = 22
        r += 1

    # Category summaries
    collab_avg = round(sum(collab_scores)/len(collab_scores), 2) if collab_scores else "—"
    culture_avg = round(sum(culture_scores)/len(culture_scores), 2) if culture_scores else "—"

    def get_lvl_label(val):
        if val == "—": return ""
        v = float(val)
        if v >= 9.0: return "Amazing!"
        if v >= 7.0: return "Good"
        if v >= 5.0: return "Keep"
        if v >= 3.0: return "Grow"
        return "Start"

    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    ws.cell(row=r, column=1, value="協作狀況（總平均）")
    ws.cell(row=r, column=4, value=collab_avg)
    ws.cell(row=r, column=5, value=get_lvl_label(collab_avg))
    style_cell_range(ws, r, 1, r, 7, font_body_bold, COLOR_WHITE, align_center)
    ws.cell(row=r, column=1).fill = PatternFill(start_color=COLOR_COLLAB_BG, end_color=COLOR_COLLAB_BG, fill_type="solid")
    ws.row_dimensions[r].height = 24
    r += 1

    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=3)
    ws.cell(row=r, column=1, value="文化實踐（總平均）")
    ws.cell(row=r, column=4, value=culture_avg)
    ws.cell(row=r, column=5, value=get_lvl_label(culture_avg))
    style_cell_range(ws, r, 1, r, 7, font_body_bold, COLOR_WHITE, align_center)
    ws.cell(row=r, column=1).fill = PatternFill(start_color=COLOR_CULTURE_BG, end_color=COLOR_CULTURE_BG, fill_type="solid")
    ws.row_dimensions[r].height = 24
    r += 1

    # 二、部屬自評細節 (文字全部公布)
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    ws.cell(row=r, column=1, value="二、部屬自評細節：部屬填答內容與自評完整實例")
    style_cell_range(ws, r, 1, r, 7, font_sub_header, COLOR_SECTION_BG, Alignment(horizontal="left", vertical="center", indent=1))
    ws.row_dimensions[r].height = 24
    r += 1

    self_details = [
        ("特質盤點", "最穩定 Top 3", ", ".join(se.get("top3_stable", [])) if (has_self and se) else "（未填）"),
        ("特質盤點", "練習中 3 項", ", ".join(se.get("top3_practice", [])) if (has_self and se) else "（未填）"),
        ("組織文化", "【信任】獨立行動與決策、主動協作、雙向溝通", se.get("values", {}).get("信任", "（未填）") if (has_self and se) else "（未填）"),
        ("組織文化", "【多元】尊重差異、多元工作方法、主動表達不同觀點", se.get("values", {}).get("多元", "（未填）") if (has_self and se) else "（未填）"),
        ("組織文化", "【實驗】透過開放心態嘗試修正與反思，勇於檢討及給予回饋", se.get("values", {}).get("實驗", "（未填）") if (has_self and se) else "（未填）"),
        ("組織文化", "【可持續】內在韌性、自我照顧、彈性的人際與工作邊界", se.get("values", {}).get("可持續", "（未填）") if (has_self and se) else "（未填）"),
    ]

    role_key = MEMBER_COMPETENCY_KEY.get(member_name, "專案經理_CGL")
    comp_titles = ROLE_COMPETENCIES_MAP.get(role_key, [])
    for c_title in comp_titles:
        c_ans = "（部屬未填寫）"
        if has_self and se.get("competencies"):
            item = next((x for x in se["competencies"] if x.get("title") == c_title or c_title in x.get("title", "")), None)
            if item and item.get("answer"): c_ans = item["answer"]
        self_details.append(("專業職能", c_title, c_ans))

    for c_sec, c_sub, c_val in self_details:
        ws.cell(row=r, column=1, value=c_sec)
        ws.cell(row=r, column=2, value=c_sub)
        ws.cell(row=r, column=3, value=c_val)
        ws.merge_cells(start_row=r, start_column=3, end_row=r, end_column=7)

        style_cell_range(ws, r, 1, r, 7, font_body, COLOR_WHITE, align_left)
        ws.cell(row=r, column=1).font = font_body_bold
        ws.cell(row=r, column=1).alignment = align_center
        ws.row_dimensions[r].height = max(30, min(120, len(c_val or "") // 35 * 18 + 18))
        r += 1

    r += 1
    # 三、質化文字回饋
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
    ws.cell(row=r, column=1, value="三、質化文字回饋：同儕文字評價與意見回饋")
    style_cell_range(ws, r, 1, r, 7, font_sub_header, COLOR_SECTION_BG, Alignment(horizontal="left", vertical="center", indent=1))
    ws.row_dimensions[r].height = 24
    r += 1

    fb_sections = [
        ("Q37. 工作與文化提升建議（改善與前進方向）", "q37_improvement_advice"),
        ("Q38. 其他補充評價與觀察", "q38_other_comments"),
        ("Q39. 肯定與感謝的話（好好星光大賞）", "q39_starlight_thanks"),
    ]

    for fb_title, fb_key in fb_sections:
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
        ws.cell(row=r, column=1, value=fb_title)
        style_cell_range(ws, r, 1, r, 7, font_table_hdr, COLOR_SECTION_BG, Alignment(horizontal="left", vertical="center", indent=1))
        ws.row_dimensions[r].height = 22
        r += 1

        if num_peers == 0:
            ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=7)
            ws.cell(row=r, column=1, value="（目前尚無同儕填答回饋）")
            style_cell_range(ws, r, 1, r, 7, font_body, COLOR_WHITE, align_left)
            ws.row_dimensions[r].height = 22
            r += 1
        else:
            for idx, r_entry in enumerate(peer_records):
                peer_label = f"同儕 {chr(65 + idx)}"
                fb_text = r_entry.get("peer_eval", {}).get(fb_key, "") or "（無填寫）"
                ws.cell(row=r, column=1, value=peer_label)
                ws.cell(row=r, column=2, value=fb_text)
                ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=7)

                style_cell_range(ws, r, 1, r, 7, font_body, COLOR_WHITE, align_left)
                ws.cell(row=r, column=1).font = font_body_bold
                ws.cell(row=r, column=1).alignment = align_center
                ws.cell(row=r, column=1).fill = PatternFill(start_color=COLOR_SECTION_BG, end_color=COLOR_SECTION_BG, fill_type="solid")
                ws.row_dimensions[r].height = max(26, min(120, len(fb_text) // 35 * 18 + 18))
                r += 1

    r += 1
    # 四、評級標準
    r = add_rating_standard_block(ws, r)

    ws.column_dimensions["A"].width = 12
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 45
    ws.column_dimensions["D"].width = 14
    ws.column_dimensions["E"].width = 14
    ws.column_dimensions["F"].width = 14
    ws.column_dimensions["G"].width = 25

# Generate Official Supervisor Workbooks Following Exact Naming Convention
# 1. 何維安 (VVN)
wb_vvn = openpyxl.Workbook()
wb_vvn.remove(wb_vvn.active)
for m in SUPERVISOR_TEAMS["何維安"]:
    add_subordinate_supervisor_sheet(wb_vvn, m)
    add_subordinate_peer_sheet(wb_vvn, m)
fname_vvn = "VVN品牌主管專用_【主管評下屬用】.xlsx"
wb_vvn.save(os.path.join(OUTPUT_DIR, fname_vvn))
print(f"Generated {fname_vvn}")

# 2. 姚品瑄
wb_ypx = openpyxl.Workbook()
wb_ypx.remove(wb_ypx.active)
for m in SUPERVISOR_TEAMS["姚品瑄"]:
    add_subordinate_supervisor_sheet(wb_ypx, m)
    add_subordinate_peer_sheet(wb_ypx, m)
fname_ypx = "姚品瑄專案主管專用_【主管評下屬用】.xlsx"
wb_ypx.save(os.path.join(OUTPUT_DIR, fname_ypx))
print(f"Generated {fname_ypx}")

# 3. 張希慈
wb_zxc = openpyxl.Workbook()
wb_zxc.remove(wb_zxc.active)
for m in SUPERVISOR_TEAMS["張希慈"]:
    add_subordinate_supervisor_sheet(wb_zxc, m)
    add_subordinate_peer_sheet(wb_zxc, m)
fname_zxc = "張希慈執行長主管專用_【主管評下屬用】.xlsx"
wb_zxc.save(os.path.join(OUTPUT_DIR, fname_zxc))
print(f"Generated {fname_zxc}")

# 4. 張希慈 (個人自評)
wb_zxc_self = openpyxl.Workbook()
wb_zxc_self.remove(wb_zxc_self.active)
add_subordinate_supervisor_sheet(wb_zxc_self, "張希慈")
add_subordinate_peer_sheet(wb_zxc_self, "張希慈")
fname_zxc_self = "張希慈個人自評_【主管評下屬用】.xlsx"
wb_zxc_self.save(os.path.join(OUTPUT_DIR, fname_zxc_self))
print(f"Generated {fname_zxc_self}")

# 5. Master Workbook with all 9 members
wb_master = openpyxl.Workbook()
wb_master.remove(wb_master.active)
for m in ALL_MEMBERS:
    add_subordinate_supervisor_sheet(wb_master, m)
    add_subordinate_peer_sheet(wb_master, m)
fname_master = "好好星球_360年中成長評估_【主管評下屬用】Master總表.xlsx"
wb_master.save(os.path.join(OUTPUT_DIR, fname_master))
print(f"Generated {fname_master}")
