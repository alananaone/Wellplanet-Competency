import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from collections import defaultdict

with open("evaluation_data.json", "r", encoding="utf-8") as f:
    entries = json.load(f)

# Colors matching unified Morandi / Pantone palette
COLOR_PINK_BLUSH = "F4CCCC"   # Top chunk header (Soft coral blush)
COLOR_PEACH_CREAM = "FCE5CD"  # Sub-table headers (Soft apricot cream)
COLOR_WHITE = "FFFFFF"
COLOR_LIGHT_ROW = "FAF7F5"    # Subtle zebra stripe
COLOR_WARN_BG = "FFF2D6"      # Pending / Missing value fill
COLOR_WARN_TEXT = "B45309"    # Warm orange-brown warning text
COLOR_SAGE_BG = "E4ECD3"      # Soft sage green for high scores / headers

# Level Colors for rating table
COLOR_L5_BG = "E4ECD3"
COLOR_L4_BG = "E2F3F0"
COLOR_L3_BG = "FFF4CD"
COLOR_L2_BG = "FCE5CD"
COLOR_L1_BG = "F4CCCC"

fill_chunk_header = PatternFill(start_color=COLOR_PINK_BLUSH, end_color=COLOR_PINK_BLUSH, fill_type="solid")
fill_sub_header = PatternFill(start_color=COLOR_PEACH_CREAM, end_color=COLOR_PEACH_CREAM, fill_type="solid")
fill_light_row = PatternFill(start_color=COLOR_LIGHT_ROW, end_color=COLOR_LIGHT_ROW, fill_type="solid")
fill_warn = PatternFill(start_color=COLOR_WARN_BG, end_color=COLOR_WARN_BG, fill_type="solid")
fill_white = PatternFill(start_color=COLOR_WHITE, end_color=COLOR_WHITE, fill_type="solid")
fill_sage = PatternFill(start_color=COLOR_SAGE_BG, end_color=COLOR_SAGE_BG, fill_type="solid")

font_chunk_title = Font(name="微軟正黑體", size=11, bold=True, color="3E2723")
font_sub_header = Font(name="微軟正黑體", size=10, bold=True, color="4E342E")
font_body = Font(name="微軟正黑體", size=9.5, color="2D2323")
font_body_bold = Font(name="微軟正黑體", size=9.5, bold=True, color="2D2323")
font_warn = Font(name="微軟正黑體", size=9.5, bold=True, color=COLOR_WARN_TEXT)
font_warn_italic = Font(name="微軟正黑體", size=9, italic=True, color=COLOR_WARN_TEXT)

thin_border = Border(
    left=Side(style='thin', color='D7CCC8'),
    right=Side(style='thin', color='D7CCC8'),
    top=Side(style='thin', color='D7CCC8'),
    bottom=Side(style='thin', color='D7CCC8')
)

align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
align_left = Alignment(horizontal="left", vertical="top", wrap_text=True)
align_header = Alignment(horizontal="center", vertical="center", wrap_text=True)

SUPERVISOR_TEAMS = {
    "張希慈": ["何維安", "陳泳璇", "張芳媐", "姚品瑄", "胡喻翔"],
    "何維安": ["林文琇"],
    "姚品瑄": ["薛筑瑄", "戴佑珍"],
    "張希慈_執行長": ["張希慈"]
}

JOB_ROLES_MAP = {
    "何維安": "品牌經理",
    "陳泳璇": "專案部門儲備主管",
    "張芳媐": "SL專案經理",
    "姚品瑄": "營運經理",
    "胡喻翔": "CGL專案經理",
    "林文琇": "視覺設計師",
    "薛筑瑄": "行政經理",
    "戴佑珍": "CGL專案經理",
    "張希慈": "執行長"
}

ROLE_COMPETENCIES = {
    "品牌經理": ["品牌定位與外部溝通一致性", "行銷策略與議題倡議", "品牌活動策劃與策展敘事", "內部品牌管理與雇主品牌", "品牌危機與聲譽風險處理", "其他創造的好事與預防的壞事"],
    "專案部門儲備主管": ["部門策略規劃與專案組合管理", "專案經理管理與培育", "部門預算與資源配置管理", "跨 LAB 專業掌握（CGL、SL、SPL）", "利害關係人管理與衝突協調", "其他創造的好事與預防的壞事"],
    "SL專案經理": ["專案企劃與現場執行", "專案時程與預算規劃管理", "需求研究與方案迭代", "外部夥伴關係經營", "社群經營與培力需求回應", "其他創造的好事與預防的壞事"],
    "營運經理": ["組織營運流程設計與優化", "制度文件與 SOP 建置", "人力資源策略與選用育留", "組織文化落實與制度轉化", "總務採購與行政庶務管理", "其他創造的好事與預防的壞事"],
    "CGL專案經理": ["專案企劃與現場執行", "專案時程與預算規劃管理", "需求研究與方案迭代", "外部夥伴關係經營", "多元教學設計與現場引導", "其他創造的好事與預防的壞事"],
    "視覺設計師": ["品牌識別系統設計與維護", "視覺設計實務", "需求釐清與創意提案", "其他創造的好事與預防的壞事"],
    "行政經理": ["人事薪資與人資系統管理", "法規與政府公文管理", "董事會與治理作業執行", "財務核銷與內控執行", "總務採購與行政庶務管理", "其他創造的好事與預防的壞事"],
    "執行長": ["策略決策與組織方向設定", "組織治理與財務風險管理", "關鍵利害關係人關係建立與維繫", "公關、演講與媒體關係", "核心團隊培養與組織文化建立", "其他創造的好事與預防的壞事"]
}

PEER_QUESTIONS = [
    ("Q24", "合作狀況", "合作狀況（溝通順暢度、資訊透明度、承諾事項達成率）", "q24_cooperation"),
    ("Q25", "工作品質", "注重細節（工作成果之品質與細緻度）", "q25_detail_oriented"),
    ("Q26", "專案進度", "準時完成（專案進度與時程掌控）", "q26_on_time"),
    ("Q27", "變動彈性", "靈活調整（面對臨時變動與突發狀況的彈性）", "q27_flexibility"),
    ("Q28", "追蹤承諾", "追蹤承諾（主動回報進度與承諾事項追蹤）", "q28_follow_up"),
    ("Q29", "透明溝通", "說明決策依據（決策與工作方法透明度）", "q29_transparency"),
    ("Q30", "多元文化", "【多元】接受不同意見（能開放聆聽反對觀點）", "q30_open_to_opposing"),
    ("Q31", "多元文化", "【多元】提出建設性觀點（在討論中給予實質建議）", "q31_constructive_opinions"),
    ("Q32", "實驗文化", "【實驗】開放調整（透過回饋持續迭代優化）", "q32_growth_mindset"),
    ("Q33", "信任文化", "【信任】分享經驗與資源（主動協助夥伴）", "q33_share_knowledge"),
    ("Q34", "可持續文化", "【可持續】讚美與肯定同儕（經常給予夥伴正向激勵）", "q34_praise_peers"),
    ("Q35", "可持續文化", "【可持續】尊重工作界線（維護彼此身心健康與邊界）", "q35_boundary_respect"),
    ("Q36", "NPS推薦", "NPS 推薦度（向他人推薦與此夥伴共事的意願）", "q36_nps_recommend"),
]

RATING_GUIDE_ROWS = [
    ("L5", "Amazing!", "遠超職位期待，表現為團隊之標竿與典範。", "讚賞其突出貢獻，探討經驗複製機制，在對話中轉為「帶其他人一起做」方向。", COLOR_L5_BG),
    ("L4", "Good", "優於職位期待，持續展現高標準成果。", "肯定並具體指出哪些行為超出標準，設定具挑戰性的下一步目標。", COLOR_L4_BG),
    ("L3", "Keep", "符合職位門檻，展現穩定的工作交付。", "確認穩定度，指出下一階可以往前之處，維持節奏並選一至兩項深化。", COLOR_L3_BG),
    ("L2", "Grow", "部分符合，部分能力/行為仍在建立階段。", "說明落差所在，聚焦一項具體可練習的行為，納入下一期 IDP 設定指標。", COLOR_L2_BG),
    ("L1", "Start", "目前的具體事證還看不到這項職能，或事證與職能要求落差明顯。", "明確對齊職位基本門檻與要求，提供即時支援與回饋引導。", COLOR_L1_BG),
]

def style_merged_range(ws, start_row, start_col, end_row, end_col, fill=None, border=thin_border):
    for r in range(start_row, end_row + 1):
        for c in range(start_col, end_col + 1):
            cell = ws.cell(r, c)
            if fill:
                cell.fill = fill
            if border:
                cell.border = border

# ==============================================================================
# 1. GENERATE ANONYMOUS PEER EVALUATION SHEET FOR EACH EMPLOYEE (逐題明細)
# ==============================================================================
def create_peer_anonymous_sheet(ws, member_name, all_entries):
    ws.views.sheetView[0].showGridLines = True
    peer_records = [e for e in all_entries if e.get("relation") == "同事" and e.get("target") == member_name]
    num_peers = len(peer_records)
    
    # Title Banner
    ws.merge_cells("A1:G1")
    t_cell = ws.cell(1, 1, value=f"好好星球文化基金會 360 年中成長評估 - 【{member_name}】同儕匿名回饋明細表")
    t_cell.font = Font(name="微軟正黑體", size=13, bold=True, color="3E2723")
    t_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    style_merged_range(ws, 1, 1, 1, 7, fill=fill_chunk_header, border=thin_border)
    ws.row_dimensions[1].height = 32

    # Sub-header notice (Full anonymity notice)
    ws.merge_cells("A2:G2")
    n_cell = ws.cell(2, 1, value=f"本報告彙整共 {num_peers} 位同儕夥伴之評估回饋（已全面匿名處理為同儕 A、同儕 B...以保護回饋者）")
    n_cell.font = Font(name="微軟正黑體", size=10, bold=True, color="4E342E")
    n_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    style_merged_range(ws, 2, 1, 2, 7, fill=fill_sub_header, border=thin_border)
    ws.row_dimensions[2].height = 24

    # Part 1 Header Banner
    ws.merge_cells("A3:G3")
    p1_cell = ws.cell(3, 1, value="一、各評估題目同儕評分明細（滿分 10 分）")
    p1_cell.font = font_sub_header
    p1_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    style_merged_range(ws, 3, 1, 3, 7, fill=fill_sub_header, border=thin_border)
    ws.row_dimensions[3].height = 24

    # Table Columns Header
    headers = ["題號", "評估面向", "評估項目與題目說明", "同儕平均得分"]
    for i in range(max(num_peers, 1)):
        headers.append(f"同儕 {chr(65+i)}")
    # Pad up to 7 columns or headers length
    while len(headers) < 7:
        headers.append("")

    for col_i, h in enumerate(headers, 1):
        c = ws.cell(4, col_i, value=h)
        c.font = font_sub_header
        c.fill = fill_sub_header
        c.alignment = align_header
        c.border = thin_border
    ws.row_dimensions[4].height = 24

    r_idx = 5
    for q_no, q_cat, q_desc, q_key in PEER_QUESTIONS:
        ws.cell(r_idx, 1, value=q_no).alignment = align_center
        ws.cell(r_idx, 2, value=q_cat).alignment = align_center
        ws.cell(r_idx, 3, value=q_desc).alignment = align_left

        scores = []
        for p_idx, p_entry in enumerate(peer_records):
            val = p_entry.get("peer_eval", {}).get(q_key)
            scores.append(val)
            c_p = ws.cell(r_idx, 5 + p_idx, value=val if val is not None else "-")
            c_p.alignment = align_center
            c_p.font = font_body
            c_p.border = thin_border

        valid_scores = [s for s in scores if s is not None]
        avg_score = (sum(valid_scores) / len(valid_scores)) if valid_scores else None
        c_avg = ws.cell(r_idx, 4, value=f"{avg_score:.1f}" if avg_score is not None else "-")
        c_avg.alignment = align_center
        c_avg.font = font_body_bold
        c_avg.border = thin_border
        c_avg.fill = fill_sage if avg_score and avg_score >= 8.5 else fill_white

        for c_i in range(1, 4):
            ws.cell(r_idx, c_i).border = thin_border
            ws.cell(r_idx, c_i).font = font_body

        ws.row_dimensions[r_idx].height = 24
        r_idx += 1

    # Part 2: Qualitative Anonymous Comments
    r_idx += 1
    ws.merge_cells(start_row=r_idx, start_column=1, end_row=r_idx, end_column=7)
    p2_cell = ws.cell(r_idx, 1, value="二、同儕質化匿名回饋彙整")
    p2_cell.font = font_sub_header
    p2_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    style_merged_range(ws, r_idx, 1, r_idx, 7, fill=fill_sub_header, border=thin_border)
    ws.row_dimensions[r_idx].height = 24
    r_idx += 1

    feedback_types = [
        ("Q37. 工作與文化提升建議（改善與前進方向）", "q37_improvement_advice"),
        ("Q38. 其他補充評價與觀察", "q38_other_comments"),
        ("Q39. 肯定與感謝的話（好好星光大賞）", "q39_starlight_thanks"),
    ]

    for f_title, f_key in feedback_types:
        ws.merge_cells(start_row=r_idx, start_column=1, end_row=r_idx, end_column=7)
        ws.cell(r_idx, 1, value=f_title).font = font_sub_header
        ws.cell(r_idx, 1).alignment = Alignment(horizontal="left", vertical="center", indent=1)
        style_merged_range(ws, r_idx, 1, r_idx, 7, fill=fill_light_row, border=thin_border)
        ws.row_dimensions[r_idx].height = 24
        r_idx += 1

        if not peer_records:
            ws.merge_cells(start_row=r_idx, start_column=1, end_row=r_idx, end_column=7)
            ws.cell(r_idx, 1, value="（目前尚無同儕填答回覆）").font = font_body
            style_merged_range(ws, r_idx, 1, r_idx, 7, fill=fill_white, border=thin_border)
            ws.row_dimensions[r_idx].height = 24
            r_idx += 1
        else:
            for p_idx, p_entry in enumerate(peer_records):
                fb_text = p_entry.get("peer_eval", {}).get(f_key, "") or "（無填寫）"
                ws.cell(r_idx, 1, value=f"同儕 {chr(65+p_idx)}").font = font_body_bold
                ws.cell(r_idx, 1).alignment = align_center
                ws.cell(r_idx, 1).fill = fill_sub_header
                ws.cell(r_idx, 1).border = thin_border

                ws.merge_cells(start_row=r_idx, start_column=2, end_row=r_idx, end_column=7)
                c_txt = ws.cell(r_idx, 2, value=fb_text)
                c_txt.font = font_body
                c_txt.alignment = Alignment(horizontal="left", vertical="center", indent=1, wrap_text=True)
                style_merged_range(ws, r_idx, 2, r_idx, 7, fill=fill_white, border=thin_border)
                
                text_len = len(str(fb_text))
                ws.row_dimensions[r_idx].height = max(24, min(100, int(text_len / 40 * 16) + 16))
                r_idx += 1
        r_idx += 1

    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 48
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 12

# ==============================================================================
# 2. GENERATE SUPERVISOR EVALUATING SUBORDINATES TABLE (主管評部屬，缺值標註)
# ==============================================================================
def create_supervisor_eval_subordinates_sheet(ws, all_entries):
    ws.views.sheetView[0].showGridLines = True

    # Title Banner
    ws.merge_cells("A1:F1")
    t_cell = ws.cell(1, 1, value="好好星球文化基金會 360 年中成長評估 - 【主管評部屬】待填缺值彙整表")
    t_cell.font = Font(name="微軟正黑體", size=13, bold=True, color="3E2723")
    t_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    style_merged_range(ws, 1, 1, 1, 6, fill=fill_chunk_header, border=thin_border)
    ws.row_dimensions[1].height = 32

    # Rating Guide Headers
    ws.merge_cells("A2:F2")
    guide_title = ws.cell(2, 1, value="職能評分等級標準與定義說明（供主管評核使用）")
    guide_title.font = Font(name="微軟正黑體", size=10.5, bold=True, color="3E2723")
    guide_title.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    style_merged_range(ws, 2, 1, 2, 6, fill=fill_sub_header, border=thin_border)
    ws.row_dimensions[2].height = 24

    ws.getCell = ws.cell
    c_g1 = ws.cell(3, 1, value="等級 (Level)")
    c_g2 = ws.cell(3, 2, value="落點名稱")
    ws.merge_cells("C3:D3")
    c_g3 = ws.cell(3, 3, value="定義說明")
    ws.merge_cells("E3:F3")
    c_g4 = ws.cell(3, 5, value="回饋語氣與後續動作")

    for cell in [c_g1, c_g2, c_g3, c_g4]:
        cell.font = font_sub_header
        cell.fill = fill_sub_header
        cell.alignment = align_header
        cell.border = thin_border
    style_merged_range(ws, 3, 3, 3, 4, fill=fill_sub_header, border=thin_border)
    style_merged_range(ws, 3, 5, 3, 6, fill=fill_sub_header, border=thin_border)
    ws.row_dimensions[3].height = 22

    for idx, (lvl, name, definition, action, bg_color) in enumerate(RATING_GUIDE_ROWS, 4):
        c1 = ws.cell(idx, 1, value=lvl)
        c2 = ws.cell(idx, 2, value=name)
        ws.merge_cells(start_row=idx, start_column=3, end_row=idx, end_column=4)
        c3 = ws.cell(idx, 3, value=definition)
        ws.merge_cells(start_row=idx, start_column=5, end_row=idx, end_column=6)
        c4 = ws.cell(idx, 5, value=action)

        row_fill = PatternFill(start_color=bg_color, end_color=bg_color, fill_type="solid")
        for cell in [c1, c2, c3, c4]:
            cell.font = font_body
            cell.fill = row_fill
            cell.border = thin_border
        c1.alignment = align_center
        c1.font = font_body_bold
        c2.alignment = align_center
        c2.font = font_body_bold
        c3.alignment = align_left
        c4.alignment = align_left
        style_merged_range(ws, idx, 3, idx, 4, fill=row_fill, border=thin_border)
        style_merged_range(ws, idx, 5, idx, 6, fill=row_fill, border=thin_border)
        ws.row_dimensions[idx].height = 26

    dv = DataValidation(type="list", formula1='"L1,L2,L3,L4,L5"', allow_blank=True)
    dv.error = '請從下拉選單中選取 L1 到 L5'
    dv.errorTitle = '評分無效'
    dv.prompt = '請選取評核等級：L1(Start), L2(Grow), L3(Keep), L4(Good), L5(Amazing!)'
    dv.promptTitle = '等級選單'
    ws.add_data_validation(dv)

    row_idx = 10
    
    # Hierarchy definition:
    hierarchy = [
        ("張希慈", "何維安", "品牌經理"),
        ("張希慈", "陳泳璇", "專案部門儲備主管"),
        ("張希慈", "張芳媐", "SL專案經理"),
        ("張希慈", "姚品瑄", "營運經理"),
        ("張希慈", "胡喻翔", "CGL專案經理"),
        ("何維安", "林文琇", "視覺設計師"),
        ("姚品瑄", "薛筑瑄", "行政經理"),
        ("姚品瑄", "戴佑珍", "CGL專案經理"),
        ("董事會", "張希慈", "執行長")
    ]

    for sup_name, sub_name, role_name in hierarchy:
        mem_entry = next((e for e in all_entries if e["target"] == sub_name and e["relation"] == "自評"), None)
        has_self = bool(mem_entry and mem_entry.get("self_eval"))

        # Chunk Top Header Banner
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=6)
        c = ws.cell(row_idx, 1, value=f"評估對象：{sub_name}（{role_name}） ｜ 直屬主管：{sup_name} ｜ 主管評核狀態：【待主管評定（缺值）】")
        c.font = font_chunk_title
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        style_merged_range(ws, row_idx, 1, row_idx, 6, fill=fill_chunk_header, border=thin_border)
        ws.row_dimensions[row_idx].height = 28
        row_idx += 1

        # Table Header
        headers = ["評估面向", "職能項目與題目說明", "部屬自評實例", "主管評核 (Lv.1-5選單)", "主管回饋與具體事證", "後續行動 / IDP 目標"]
        for col_i, h in enumerate(headers, 1):
            cell = ws.cell(row_idx, col_i, value=h)
            cell.font = font_sub_header
            cell.fill = fill_sub_header
            cell.alignment = align_header
            cell.border = thin_border
        ws.row_dimensions[row_idx].height = 24
        row_idx += 1

        # 1. 組織文化實踐 (4 rows)
        cult_start = row_idx
        culture_items = [
            ("【信任】獨立行動與決策、主動協作、雙向溝通", mem_entry.get("self_eval", {}).get("values", {}).get("信任") if has_self else None),
            ("【多元】尊重差異、多元工作方法、主動表達不同觀點", mem_entry.get("self_eval", {}).get("values", {}).get("多元") if has_self else None),
            ("【實驗】透過開放心態嘗試修正與反思，勇於檢討給予回饋", mem_entry.get("self_eval", {}).get("values", {}).get("實驗") if has_self else None),
            ("【可持續】內在韌性、自我照顧、彈性的人際與工作邊界", mem_entry.get("self_eval", {}).get("values", {}).get("可持續") if has_self else None),
        ]

        for c_title, c_self in culture_items:
            ws.cell(row_idx, 2, value=c_title).font = font_body
            ws.cell(row_idx, 2).alignment = align_left
            ws.cell(row_idx, 2).border = thin_border

            c3 = ws.cell(row_idx, 3, value=c_self if c_self else "（部屬未填寫自評實例）")
            c3.font = font_body if c_self else font_warn_italic
            c3.alignment = align_left
            c3.border = thin_border
            if not c_self: c3.fill = fill_warn

            ws.row_dimensions[row_idx].height = 36
            row_idx += 1
        cult_end = row_idx - 1

        # Merge Col A for 組織文化
        ws.merge_cells(start_row=cult_start, start_column=1, end_row=cult_end, end_column=1)
        c_cult = ws.cell(cult_start, 1, value="組織文化\n(整體回饋)")
        c_cult.font = font_body_bold
        c_cult.alignment = align_center
        style_merged_range(ws, cult_start, 1, cult_end, 1, fill=fill_white, border=thin_border)

        # Merge Col D for 文化 (不適用)
        ws.merge_cells(start_row=cult_start, start_column=4, end_row=cult_end, end_column=4)
        c_lv_na = ws.cell(cult_start, 4, value="不適用\n(以文字回饋)")
        c_lv_na.font = font_warn_italic
        c_lv_na.alignment = align_center
        style_merged_range(ws, cult_start, 4, cult_end, 4, fill=fill_warn, border=thin_border)

        # Merge Col E for 主管文化回饋 (缺值標註)
        ws.merge_cells(start_row=cult_start, start_column=5, end_row=cult_end, end_column=5)
        c_fb = ws.cell(cult_start, 5, value="【待主管填寫文化實踐回饋】")
        c_fb.font = font_warn
        c_fb.alignment = align_left
        style_merged_range(ws, cult_start, 5, cult_end, 5, fill=fill_warn, border=thin_border)

        # Merge Col F for 後續動作 (缺值標註)
        ws.merge_cells(start_row=cult_start, start_column=6, end_row=cult_end, end_column=6)
        c_act = ws.cell(cult_start, 6, value="【待主管設定文化成長動作】")
        c_act.font = font_warn
        c_act.alignment = align_left
        style_merged_range(ws, cult_start, 6, cult_end, 6, fill=fill_warn, border=thin_border)

        # 2. 專業職能 (6 rows)
        comp_titles = ROLE_COMPETENCIES.get(role_name, [])
        comp_start = row_idx
        for c_t in comp_titles:
            ws.cell(row_idx, 2, value=c_t).font = font_body
            ws.cell(row_idx, 2).alignment = align_left
            ws.cell(row_idx, 2).border = thin_border

            # Subordinate answer
            self_ans = None
            if has_self and mem_entry["self_eval"].get("competencies"):
                found_c = next((item["answer"] for item in mem_entry["self_eval"]["competencies"] if item["title"] == c_t), None)
                if found_c: self_ans = found_c

            c3 = ws.cell(row_idx, 3, value=self_ans if self_ans else "（部屬未填寫自評實例）")
            c3.font = font_body if self_ans else font_warn_italic
            c3.alignment = align_left
            c3.border = thin_border
            if not self_ans: c3.fill = fill_warn

            # Col D: Lv Dropdown (Missing)
            c4 = ws.cell(row_idx, 4, value="")
            c4.font = font_body_bold
            c4.alignment = align_center
            c4.border = thin_border
            c4.fill = fill_warn
            dv.add(c4)

            # Col E: Supervisor Feedback (Missing)
            c5 = ws.cell(row_idx, 5, value="【待主管填寫評核回饋】")
            c5.font = font_warn
            c5.alignment = align_left
            c5.border = thin_border
            c5.fill = fill_warn

            # Col F: IDP (Missing)
            c6 = ws.cell(row_idx, 6, value="【待設定下一步目標】")
            c6.font = font_warn
            c6.alignment = align_left
            c6.border = thin_border
            c6.fill = fill_warn

            ws.row_dimensions[row_idx].height = 36
            row_idx += 1
        comp_end = row_idx - 1

        # Merge Col A for 專業職能
        ws.merge_cells(start_row=comp_start, start_column=1, end_row=comp_end, end_column=1)
        c_comp = ws.cell(comp_start, 1, value="專業職能")
        c_comp.font = font_body_bold
        c_comp.alignment = align_center
        style_merged_range(ws, comp_start, 1, comp_end, 1, fill=fill_white, border=thin_border)

        row_idx += 2

    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 34
    ws.column_dimensions['C'].width = 50
    ws.column_dimensions['D'].width = 18
    ws.column_dimensions['E'].width = 36
    ws.column_dimensions['F'].width = 30

# ==============================================================================
# 3. BUILD AND SAVE ALL WORKBOOKS
# ==============================================================================
print("Generating Master Workbooks...")

# Master Workbook with all sheets
wb_master = openpyxl.load_workbook("好好星球_360年中成長評估_主管分流與完整彙整表.xlsx")

# Add "主管評部屬待填總表"
if "主管評部屬_待填彙整表" in wb_master.sheetnames:
    del wb_master["主管評部屬_待填彙整表"]
ws_sup_eval = wb_master.create_sheet(title="主管評部屬_待填彙整表")
create_supervisor_eval_subordinates_sheet(ws_sup_eval, entries)

# Add Anonymous Peer Sheets for each employee
all_member_names = ["何維安", "姚品瑄", "張芳媐", "戴佑珍", "林文琇", "胡喻翔", "薛筑瑄", "陳泳璇", "張希慈"]
for m_name in all_member_names:
    s_title = f"同儕匿名_{m_name}"
    if s_title in wb_master.sheetnames:
        del wb_master[s_title]
    ws_peer_anon = wb_master.create_sheet(title=s_title)
    create_peer_anonymous_sheet(ws_peer_anon, m_name, entries)

wb_master.save("好好星球_360年中成長評估_主管分流與完整彙整表.xlsx")
print("Saved 好好星球_360年中成長評估_主管分流與完整彙整表.xlsx with all sheets!")

# Dedicated Standalone Workbook: 同儕匿名回饋總表(每人獨立Sheet)
wb_peers_all = openpyxl.Workbook()
wb_peers_all.remove(wb_peers_all.active)
for m_name in all_member_names:
    ws_m = wb_peers_all.create_sheet(title=f"【{m_name}】同儕匿名回饋")
    create_peer_anonymous_sheet(ws_m, m_name, entries)
wb_peers_all.save("同儕匿名回饋_全體員工每人獨立Sheet總表.xlsx")
print("Saved 同儕匿名回饋_全體員工每人獨立Sheet總表.xlsx!")

# Dedicated Standalone Workbook: 主管評部屬_待填缺值彙整表
wb_sup_eval_only = openpyxl.Workbook()
ws_seo = wb_sup_eval_only.active
ws_seo.title = "主管評部屬_待填評核表"
create_supervisor_eval_subordinates_sheet(ws_seo, entries)
wb_sup_eval_only.save("主管評部屬_待填缺值彙整表.xlsx")
print("Saved 主管評部屬_待填缺值彙整表.xlsx!")

