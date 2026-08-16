import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

with open("evaluation_data.json", "r", encoding="utf-8") as f:
    entries = json.load(f)

# Styles
font_title = Font(name="微軟正黑體", size=14, bold=True, color="1E293B")
font_section = Font(name="微軟正黑體", size=11, bold=True, color="FFFFFF")
font_header = Font(name="微軟正黑體", size=10, bold=True, color="1E293B")
font_body = Font(name="微軟正黑體", size=10, color="334155")
font_sub = Font(name="微軟正黑體", size=9, italic=True, color="64748B")

fill_section_indigo = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
fill_header_gray = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
fill_alert_yellow = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")

thin_border = Border(
    left=Side(style='thin', color='CBD5E1'),
    right=Side(style='thin', color='CBD5E1'),
    top=Side(style='thin', color='CBD5E1'),
    bottom=Side(style='thin', color='CBD5E1')
)

align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
align_left = Alignment(horizontal="left", vertical="top", wrap_text=True)
align_header = Alignment(horizontal="center", vertical="center", wrap_text=True)

def create_supervisor_workbook(supervisor_name, member_names, filename):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "部屬自評詳細表"
    ws.views.sheetView[0].showGridLines = True
    
    # Title Block
    ws.cell(row=1, column=1, value=f"好好星球文化基金會 360 年中成長評估 - 【{supervisor_name}】部屬自評專用表").font = font_title
    ws.row_dimensions[1].height = 30
    
    row_idx = 3
    for member in member_names:
        mem_entry = next((e for e in entries if e["target"] == member and e["relation"] == "自評"), None)
        
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=4)
        c = ws.cell(row=row_idx, column=1, value=f"👤 部屬姓名：{member}" + (f" （職位：{mem_entry['self_eval']['job_role']}）" if mem_entry and mem_entry.get('self_eval') else " （⚠️ 尚未填寫自評表單）"))
        c.font = font_section
        c.fill = fill_section_indigo
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[row_idx].height = 26
        row_idx += 1
        
        if not mem_entry or not mem_entry.get("self_eval"):
            ws.cell(row=row_idx, column=1, value="狀態說明").font = font_header
            ws.cell(row=row_idx, column=1).fill = fill_header_gray
            ws.cell(row=row_idx, column=1).border = thin_border
            
            ws.merge_cells(start_row=row_idx, start_column=2, end_row=row_idx, end_column=4)
            c2 = ws.cell(row=row_idx, column=2, value=f"目前表單中尚未收到 {member} 的自我評估回覆紀錄。")
            c2.font = font_body
            c2.fill = fill_alert_yellow
            c2.border = thin_border
            ws.row_dimensions[row_idx].height = 24
            row_idx += 2
            continue
            
        se = mem_entry["self_eval"]
        
        headers = ["評估面向 / 題組", "題目標題與說明", "部屬自評回覆內容", "主管評核與回饋備註"]
        for col_i, h in enumerate(headers, 1):
            cell = ws.cell(row=row_idx, column=col_i, value=h)
            cell.font = font_header
            cell.fill = fill_header_gray
            cell.alignment = align_header
            cell.border = thin_border
        ws.row_dimensions[row_idx].height = 24
        row_idx += 1
        
        qa_list = [
            ("基本資訊", "填答時間 / 電子郵件", f"{mem_entry['timestamp']} / {mem_entry['email']}"),
            ("基本資訊", "評估職位", se.get("job_role", "未填寫")),
            ("工作特質盤點", "過去展現最穩定、最具代表性的特質 Top 3", "、".join(se.get("top3_stable", []))),
            ("工作特質盤點", "目前還在練習、或希望未來更穩定發展的特質 3項", "、".join(se.get("top3_practice", []))),
            ("四大文化實踐", "【信任】能獨立行動與決策、主動協作、雙向溝通，在高彈性下給予彼此信任", se.get("values", {}).get("信任", "（無填寫）")),
            ("四大文化實踐", "【多元】尊重差異、多元工作方法、主動表達不同觀點與想法", se.get("values", {}).get("多元", "（無填寫）")),
            ("四大文化實踐", "【實驗】透過開放的心態不斷嘗試、修正與反思，勇於檢討及給予回饋", se.get("values", {}).get("實驗", "（無填寫）")),
            ("四大文化實踐", "【可持續】內在韌性、自我照顧、彈性的人際與工作邊界", se.get("values", {}).get("可持續", "（無填寫）")),
        ]
        
        for comp in se.get("competencies", []):
            qa_list.append((f"職能展現 ({se.get('job_role')})", comp["title"], comp["answer"] or "（無填寫）"))
            
        for q_title, q_ans in se.get("reflection", {}).items():
            qa_list.append(("組織卡關點與展望", q_title, q_ans or "（無填寫）"))
            
        for category, q_title, q_ans in qa_list:
            c1 = ws.cell(row=row_idx, column=1, value=category)
            c2 = ws.cell(row=row_idx, column=2, value=q_title)
            c3 = ws.cell(row=row_idx, column=3, value=q_ans)
            c4 = ws.cell(row=row_idx, column=4, value="")
            
            for c in [c1, c2, c3, c4]:
                c.border = thin_border
                c.font = font_body
                c.alignment = align_left
            c1.alignment = align_center
            
            text_len = len(str(q_ans))
            ws.row_dimensions[row_idx].height = max(24, min(140, int(text_len / 40 * 18) + 18))
            row_idx += 1
            
        row_idx += 2
        
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 36
    ws.column_dimensions['C'].width = 68
    ws.column_dimensions['D'].width = 30
    
    wb.save(filename)
    print(f"Generated {filename}")

create_supervisor_workbook("張希慈", ["何維安", "陳泳璇", "張芳媐", "姚品瑄", "胡喻翔"], "自評_張希慈主管專用_部屬自評彙整表.xlsx")
create_supervisor_workbook("何維安", ["林文琇"], "自評_何維安主管專用_部屬自評彙整表.xlsx")
create_supervisor_workbook("姚品瑄", ["戴佑珍", "薛筑瑄"], "自評_姚品瑄主管專用_部屬自評彙整表.xlsx")
create_supervisor_workbook("張希慈(執行長自評)", ["張希慈"], "自評_張希慈執行長自評表.xlsx")
