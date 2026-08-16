import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

with open("evaluation_data.json", "r", encoding="utf-8") as f:
    entries = json.load(f)

# Styles
font_title = Font(name="微軟正黑體", size=14, bold=True, color="1E293B")
font_section = Font(name="微軟正黑體", size=11, bold=True, color="FFFFFF")
font_header = Font(name="微軟正黑體", size=10, bold=True, color="1E293B")
font_body = Font(name="微軟正黑體", size=10, color="334155")
font_sub = Font(name="微軟正黑體", size=9, italic=True, color="64748B")
font_tag_green = Font(name="微軟正黑體", size=9, bold=True, color="166534")
font_tag_orange = Font(name="微軟正黑體", size=9, bold=True, color="9A3412")

fill_section_blue = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
fill_section_indigo = PatternFill(start_color="4F46E5", end_color="4F46E5", fill_type="solid")
fill_section_emerald = PatternFill(start_color="059669", end_color="059669", fill_type="solid")
fill_section_purple = PatternFill(start_color="7C3AED", end_color="7C3AED", fill_type="solid")
fill_header_gray = PatternFill(start_color="F1F5F9", end_color="F1F5F9", fill_type="solid")
fill_card_bg = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")
fill_alert_yellow = PatternFill(start_color="FEF3C7", end_color="FEF3C7", fill_type="solid")

thin_border = Border(
    left=Side(style='thin', color='CBD5E1'),
    right=Side(style='thin', color='CBD5E1'),
    top=Side(style='thin', color='CBD5E1'),
    bottom=Side(style='thin', color='CBD5E1')
)

align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
align_left = Alignment(horizontal="left", vertical="top", wrap_text=True)
align_header = Alignment(horizontal="center", vertical="center", wrap_text=True)

wb = openpyxl.Workbook()
# remove default sheet
wb.remove(wb.active)

# Helper function to create a Self-Eval sheet for a supervisor
def build_self_eval_sheet(wb, sheet_title, supervisor_name, member_names, self_entries):
    ws = wb.create_sheet(title=sheet_title[:31])
    ws.views.sheetView[0].showGridLines = True
    
    # Title Block
    ws.cell(row=1, column=1, value=f"好好星球文化基金會 360 年中成長評估 - 【{supervisor_name}】部屬自評彙整表").font = font_title
    ws.row_dimensions[1].height = 30
    
    row_idx = 3
    
    # Filter matching members
    for member in member_names:
        # Find self eval entry
        mem_entry = next((e for e in self_entries if e["target"] == member and e["relation"] == "自評"), None)
        
        # Section Header for Member
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=4)
        c = ws.cell(row=row_idx, column=1, value=f"👤 部屬姓名：{member}" + (f" （職位：{mem_entry['self_eval']['job_role']}）" if mem_entry and mem_entry.get('self_eval') else " （⚠️ 尚未填寫自評表單）"))
        c.font = font_section
        c.fill = fill_section_indigo
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[row_idx].height = 26
        row_idx += 1
        
        if not mem_entry or not mem_entry.get("self_eval"):
            ws.cell(row=row_idx, column=1, value="狀態說明").font = font_header
            ws.cell(row=row_idx, column=1).fill = fill_header_gray
            ws.cell(row=row_idx, column=1).border = thin_border
            
            ws.merge_cells(start_row=row_idx, start_column=2, end_row=row_idx, end_column=4)
            c2 = ws.cell(row=row_idx, column=2, value=f"目前表單中尚未收到 {member} 的自我評估回覆紀錄。")
            c2.font = font_body
            c2.fill = fill_alert_yellow
            c2.border = thin_border
            ws.row_dimensions[row_idx].height = 24
            row_idx += 2
            continue
            
        se = mem_entry["self_eval"]
        
        # Table of Questions & Answers
        headers = ["評估面向 / 題組", "題目標題與說明", "部屬自評回覆內容", "主管評核與回饋備註"]
        for col_i, h in enumerate(headers, 1):
            cell = ws.cell(row=row_idx, column=col_i, value=h)
            cell.font = font_header
            cell.fill = fill_header_gray
            cell.alignment = align_header
            cell.border = thin_border
        ws.row_dimensions[row_idx].height = 24
        row_idx += 1
        
        # 1. 基本資訊
        qa_list = [
            ("基本資訊", "填答時間 / 電子郵件", f"{mem_entry['timestamp']} / {mem_entry['email']}"),
            ("基本資訊", "評估職位", se.get("job_role", "未填寫")),
            ("工作特質盤點", "過去展現最穩定、最具代表性的特質 Top 3", "、".join(se.get("top3_stable", []))),
            ("工作特質盤點", "目前還在練習、或希望未來更穩定發展的特質 3 項", "、".join(se.get("top3_practice", []))),
            ("四大文化實踐", "【信任】能獨立行動與決策、主動協作、雙向溝通，在高彈性下給予彼此信任", se.get("values", {}).get("信任", "（無填寫）")),
            ("四大文化實踐", "【多元】尊重差異、多元工作方法、主動表達不同觀點與想法", se.get("values", {}).get("多元", "（無填寫）")),
            ("四大文化實踐", "【實驗】透過開放的心態不斷嘗試、修正與反思，勇於檢討及給予回饋", se.get("values", {}).get("實驗", "（無填寫）")),
            ("四大文化實踐", "【可持續】內在韌性、自我照顧、彈性的人際與工作邊界", se.get("values", {}).get("可持續", "（無填寫）")),
        ]
        
        # 2. 職能展現
        for comp in se.get("competencies", []):
            qa_list.append((f"職能展現 ({se.get('job_role')})", comp["title"], comp["answer"] or "（無填寫）"))
            
        # 3. 組織卡關點與未來展望
        for q_title, q_ans in se.get("reflection", {}).items():
            qa_list.append(("組織卡關點與展望", q_title, q_ans or "（無填寫）"))
            
        for category, q_title, q_ans in qa_list:
            c1 = ws.cell(row=row_idx, column=1, value=category)
            c2 = ws.cell(row=row_idx, column=2, value=q_title)
            c3 = ws.cell(row=row_idx, column=3, value=q_ans)
            c4 = ws.cell(row=row_idx, column=4, value="") # for manager feedback
            
            for c in [c1, c2, c3, c4]:
                c.border = thin_border
                c.font = font_body
                c.alignment = align_left
            c1.alignment = align_center
            
            # calculate height
            text_len = len(str(q_ans))
            ws.row_dimensions[row_idx].height = max(24, min(120, int(text_len / 40 * 18) + 18))
            row_idx += 1
            
        row_idx += 2
        
    # Column widths
    ws.column_dimensions['A'].width = 18
    ws.column_dimensions['B'].width = 36
    ws.column_dimensions['C'].width = 65
    ws.column_dimensions['D'].width = 30

# Build Self Eval sheets for each supervisor
build_self_eval_sheet(wb, "自評_張希慈的主管包", "張希慈", ["何維安", "陳泳璇", "張芳媐", "姚品瑄"], entries)
build_self_eval_sheet(wb, "自評_何維安的主管包", "何維安", ["林文琇"], entries)
build_self_eval_sheet(wb, "自評_姚品瑄的主管包", "姚品瑄", ["薛筑瑄", "戴佑珍"], entries)
build_self_eval_sheet(wb, "自評_張希慈(執行長)", "張希慈", ["張希慈"], entries)

# Build "評主管總表" Sheet
ws_sup = wb.create_sheet(title="評主管總表")
ws_sup.views.sheetView[0].showGridLines = True
ws_sup.cell(row=1, column=1, value="好好星球文化基金會 360 年中成長評估 - 【評主管】回覆總表").font = font_title
ws_sup.row_dimensions[1].height = 30

sup_headers = [
    "時間戳記", "填答者Email", "受評主管", 
    "Q4.尋求協助容易度", "Q5.具體引導頻率", "Q6.表現改善幅度", "Q7.跨部門協助推動",
    "Q8.資源評估周全度", "Q9.失誤建設性回應", "Q10.肯定與認可頻率", "Q11.整體工作成效",
    "Q12.【信任】表達真實想法", "Q13.【多元】決策前聆聽意見", "Q14.【實驗】嘗試新方式承擔風險",
    "Q15.【實驗】試錯空間與支持", "Q16.【可持續】讚美肯定夥伴", "Q17.【可持續】尊重個人界線",
    "Q18.主管NPS推薦度(0-10)", "Q19.整體滿意度",
    "Q20.願景與使命理解引導(質化)", "Q21.管理與文化精神提升建議(質化)", "Q22.其他補充評價(質化)", "Q23.肯定與感謝的話(星光大賞)"
]

for col_i, h in enumerate(sup_headers, 1):
    c = ws_sup.cell(row=2, column=col_i, value=h)
    c.font = font_header
    c.fill = fill_section_blue
    c.font = Font(name="微軟正黑體", size=9, bold=True, color="FFFFFF")
    c.alignment = align_header
    c.border = thin_border
ws_sup.row_dimensions[2].height = 28

row_idx = 3
for e in entries:
    if e["relation"] == "主管":
        se = e.get("supervisor_eval", {})
        row_vals = [
            e["timestamp"], e["email"], e["target"],
            se.get("q4_help_easy"), se.get("q5_guidance_freq"), se.get("q6_improve_degree"), se.get("q7_cross_dept"),
            se.get("q8_resource_eval"), se.get("q9_constructive_mistake"), se.get("q10_recognition"), se.get("q11_overall_performance"),
            se.get("q12_trust_express"), se.get("q13_diversity_listen"), se.get("q14_experiment_try"),
            se.get("q15_experiment_psych_safety"), se.get("q16_sustain_praise"), se.get("q17_sustain_boundary"),
            se.get("q18_nps_recommend"), se.get("q19_satisfaction"),
            se.get("q20_vision_mission"), se.get("q21_improvement_advice"), se.get("q22_other_comments"), se.get("q23_starlight_thanks")
        ]
        for col_i, val in enumerate(row_vals, 1):
            c = ws_sup.cell(row=row_idx, column=col_i, value=val)
            c.font = font_body
            c.border = thin_border
            c.alignment = align_center if col_i <= 19 else align_left
        ws_sup.row_dimensions[row_idx].height = 45
        row_idx += 1

for c in range(1, len(sup_headers) + 1):
    ws_sup.column_dimensions[get_column_letter(c)].width = 16 if c <= 19 else 35

# Build "評同事總表" Sheet
ws_peer = wb.create_sheet(title="評同事總表")
ws_peer.views.sheetView[0].showGridLines = True
ws_peer.cell(row=1, column=1, value="好好星球文化基金會 360 年中成長評估 - 【評同事】回覆總表").font = font_title
ws_peer.row_dimensions[1].height = 30

peer_headers = [
    "時間戳記", "填答者Email", "受評同事",
    "Q24.合作狀況", "Q25.注重細節", "Q26.準時完成工作", "Q27.靈活調整協調", "Q28.承擔追蹤承諾", "Q29.說明想法決策依據",
    "Q30.【多元】接受不同意見", "Q31.【多元】表達建設性觀點", "Q32.【實驗】面對建議開放調整",
    "Q33.【信任】主動分享經驗資源", "Q34.【可持續】讚美肯定同事", "Q35.【可持續】尊重個人界線",
    "Q36.同事NPS推薦度(0-10)",
    "Q37.提升工作與文化表現建議(質化)", "Q38.其他補充評價(質化)", "Q39.肯定與感謝的話(星光大賞)"
]

for col_i, h in enumerate(peer_headers, 1):
    c = ws_peer.cell(row=2, column=col_i, value=h)
    c.font = Font(name="微軟正黑體", size=9, bold=True, color="FFFFFF")
    c.fill = fill_section_emerald
    c.alignment = align_header
    c.border = thin_border
ws_peer.row_dimensions[2].height = 28

row_idx = 3
for e in entries:
    if e["relation"] == "同事":
        pe = e.get("peer_eval", {})
        row_vals = [
            e["timestamp"], e["email"], e["target"],
            pe.get("q24_cooperation"), pe.get("q25_detail_oriented"), pe.get("q26_on_time"), pe.get("q27_flexibility"),
            pe.get("q28_follow_up"), pe.get("q29_transparency"),
            pe.get("q30_open_to_opposing"), pe.get("q31_constructive_opinions"), pe.get("q32_growth_mindset"),
            pe.get("q33_share_knowledge"), pe.get("q34_praise_peers"), pe.get("q35_boundary_respect"),
            pe.get("q36_nps_recommend"),
            pe.get("q37_improvement_advice"), pe.get("q38_other_comments"), pe.get("q39_starlight_thanks")
        ]
        for col_i, val in enumerate(row_vals, 1):
            c = ws_peer.cell(row=row_idx, column=col_i, value=val)
            c.font = font_body
            c.border = thin_border
            c.alignment = align_center if col_i <= 16 else align_left
        ws_peer.row_dimensions[row_idx].height = 40
        row_idx += 1

for c in range(1, len(peer_headers) + 1):
    ws_peer.column_dimensions[get_column_letter(c)].width = 16 if c <= 16 else 35

output_filename = "好好星球_360年中成長評估_主管分流與完整彙整表.xlsx"
wb.save(output_filename)
print(f"Generated {output_filename} successfully!")
