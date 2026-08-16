import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

with open("evaluation_data.json", "r", encoding="utf-8") as f:
    entries = json.load(f)

# Colors matching user's template (#F4CCCC, #FCE5CD) and warm aesthetic
COLOR_PINK_BLUSH = "F4CCCC"   # Top chunk header
COLOR_PEACH_CREAM = "FCE5CD"  # Sub-table headers
COLOR_WHITE = "FFFFFF"
COLOR_LIGHT_ROW = "FAF7F5"    # Subtle zebra stripe
COLOR_WARN_BG = "FFF2D6"      # Pending self-eval

fill_chunk_header = PatternFill(start_color=COLOR_PINK_BLUSH, end_color=COLOR_PINK_BLUSH, fill_type="solid")
fill_sub_header = PatternFill(start_color=COLOR_PEACH_CREAM, end_color=COLOR_PEACH_CREAM, fill_type="solid")
fill_light_row = PatternFill(start_color=COLOR_LIGHT_ROW, end_color=COLOR_LIGHT_ROW, fill_type="solid")
fill_warn = PatternFill(start_color=COLOR_WARN_BG, end_color=COLOR_WARN_BG, fill_type="solid")
fill_white = PatternFill(start_color=COLOR_WHITE, end_color=COLOR_WHITE, fill_type="solid")

font_chunk_title = Font(name="微軟正黑體", size=11, bold=True, color="3E2723")
font_sub_header = Font(name="微軟正黑體", size=10, bold=True, color="4E342E")
font_body = Font(name="微軟正黑體", size=9.5, color="2D2323")
font_body_bold = Font(name="微軟正黑體", size=9.5, bold=True, color="2D2323")
font_sub_note = Font(name="微軟正黑體", size=9, italic=True, color="795548")
font_warn = Font(name="微軟正黑體", size=9.5, bold=True, color="B45309")

thin_border = Border(
    left=Side(style='thin', color='D7CCC8'),
    right=Side(style='thin', color='D7CCC8'),
    top=Side(style='thin', color='D7CCC8'),
    bottom=Side(style='thin', color='D7CCC8')
)

thick_bottom = Border(
    left=Side(style='thin', color='D7CCC8'),
    right=Side(style='thin', color='D7CCC8'),
    top=Side(style='thin', color='D7CCC8'),
    bottom=Side(style='medium', color='8D6E63')
)

align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
align_left = Alignment(horizontal="left", vertical="top", wrap_text=True)
align_header = Alignment(horizontal="center", vertical="center", wrap_text=True)

SUPERVISOR_TEAMS = {
    "張希慈": ["何維安", "陳泳璇", "張芳媐", "姚品瑄", "胡喻翔"],
    "何維安": ["林文琇"],
    "姚品瑄": ["薛筑瑄", "戴佑珍"],
    "張希慈_執行長": ["張希慈"]
}

def create_supervisor_chunk_sheet(ws, supervisor_name, member_names, all_entries):
    ws.views.sheetView[0].showGridLines = True
    
    # Title Banner
    ws.merge_cells("A1:E1")
    t_cell = ws.cell(1, 1, value=f"好好星球文化基金會 360 年中成長評估 - 【{supervisor_name}】部屬自評與主管評核表")
    t_cell.font = Font(name="微軟正黑體", size=13, bold=True, color="3E2723")
    t_cell.fill = fill_chunk_header
    t_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[1].height = 32
    
    ws.merge_cells("A2:E2")
    sub_cell = ws.cell(2, 1, value=f"評分標準說明：Lv.1 ~ Lv.5 （符合期待 ～ 超越期待） ｜ 填寫方式：可在每題右方直接給予等級評分與 Feedback 回饋")
    sub_cell.font = font_sub_note
    sub_cell.fill = PatternFill(start_color="FFF9F5", end_color="FFF9F5", fill_type="solid")
    sub_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    ws.row_dimensions[2].height = 20

    row_idx = 4
    for member in member_names:
        mem_entry = next((e for e in all_entries if e["target"] == member and e["relation"] == "自評"), None)
        
        # 1. Chunk Top Header
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=5)
        role_str = mem_entry['self_eval']['job_role'] if mem_entry and mem_entry.get('self_eval') else "（⚠️ 尚未填寫自評）"
        time_str = f" | 填答時間：{mem_entry['timestamp']}" if mem_entry else ""
        c = ws.cell(row_idx, 1, value=f"👤 部屬姓名：{member}    ｜    職位：{role_str}{time_str}")
        c.font = font_chunk_title
        c.fill = fill_chunk_header
        c.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        ws.row_dimensions[row_idx].height = 28
        row_idx += 1
        
        if not mem_entry or not mem_entry.get("self_eval"):
            ws.cell(row_idx, 1, value="自評狀態").font = font_body_bold
            ws.cell(row_idx, 1).fill = fill_sub_header
            ws.cell(row_idx, 1).border = thin_border
            ws.cell(row_idx, 1).alignment = align_center
            
            ws.merge_cells(start_row=row_idx, start_column=2, end_row=row_idx, end_column=5)
            c2 = ws.cell(row_idx, 2, value=f"目前表單中尚未收到 {member} 的自我評估回覆紀錄。收到新回覆後可重新載入。")
            c2.font = font_warn
            c2.fill = fill_warn
            c2.border = thin_border
            c2.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            ws.row_dimensions[row_idx].height = 26
            row_idx += 3
            continue
            
        se = mem_entry["self_eval"]
        
        # 2. 工作特質盤點 Chunk
        ws.cell(row_idx, 1, value="工作特質盤點").font = font_body_bold
        ws.cell(row_idx, 1).fill = fill_sub_header
        ws.cell(row_idx, 1).border = thin_border
        ws.cell(row_idx, 1).alignment = align_center
        
        traits_text = f"【最穩定代表 Top 3】：" + ("、".join(se.get("top3_stable", [])) or "無") + "\n" + \
                      f"【練習中/期望發展】：" + ("、".join(se.get("top3_practice", [])) or "無")
        ws.merge_cells(start_row=row_idx, start_column=2, end_row=row_idx, end_column=5)
        c_trait = ws.cell(row_idx, 2, value=traits_text)
        c_trait.font = font_body
        c_trait.fill = fill_white
        c_trait.border = thin_border
        c_trait.alignment = align_left
        ws.row_dimensions[row_idx].height = 36
        row_idx += 1
        
        # 3. 四大文化實踐 Chunk Table
        headers = ["評估面向", "四大文化定義與說明", "部屬自評實例 (STAR)", "Lv.", "主管 Feedback"]
        for col_i, h in enumerate(headers, 1):
            cell = ws.cell(row_idx, col_i, value=h)
            cell.font = font_sub_header
            cell.fill = fill_sub_header
            cell.alignment = align_header
            cell.border = thin_border
        ws.row_dimensions[row_idx].height = 24
        row_idx += 1
        
        culture_rows = [
            ("組織文化", "【信任】獨立行動與決策、主動協作、雙向溝通，在高彈性下給予彼此信任", se.get("values", {}).get("信任", "（無）")),
            ("組織文化", "【多元】尊重差異、多元工作方法、主動表達不同觀點與想法", se.get("values", {}).get("多元", "（無）")),
            ("組織文化", "【實驗】透過開放心態嘗試修正與反思，勇於檢討及給予回饋", se.get("values", {}).get("實驗", "（無）")),
            ("組織文化", "【可持續】內在韌性、自我照顧、彈性的人際與工作邊界", se.get("values", {}).get("可持續", "（無）")),
        ]
        
        for cat, title, ans in culture_rows:
            c1 = ws.cell(row_idx, 1, value=cat)
            c2 = ws.cell(row_idx, 2, value=title)
            c3 = ws.cell(row_idx, 3, value=ans)
            c4 = ws.cell(row_idx, 4, value="") # Lv.
            c5 = ws.cell(row_idx, 5, value="") # Feedback
            
            for cell in [c1, c2, c3, c4, c5]:
                cell.border = thin_border
                cell.font = font_body
                cell.alignment = align_left
            c1.alignment = align_center
            c4.alignment = align_center
            
            text_len = len(str(ans))
            ws.row_dimensions[row_idx].height = max(26, min(120, int(text_len / 40 * 18) + 18))
            row_idx += 1
            
        # 4. 職能展現 Chunk Table
        if se.get("competencies"):
            ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=5)
            comp_banner = ws.cell(row_idx, 1, value=f"⭐ 【{se.get('job_role')}】專業職能展現實例")
            comp_banner.font = Font(name="微軟正黑體", size=10, bold=True, color="4E342E")
            comp_banner.fill = fill_sub_header
            comp_banner.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            ws.row_dimensions[row_idx].height = 24
            row_idx += 1
            
            for comp in se["competencies"]:
                c1 = ws.cell(row_idx, 1, value="專業職能")
                c2 = ws.cell(row_idx, 2, value=comp["title"])
                c3 = ws.cell(row_idx, 3, value=comp["answer"] or "（無填寫）")
                c4 = ws.cell(row_idx, 4, value="")
                c5 = ws.cell(row_idx, 5, value="")
                
                for cell in [c1, c2, c3, c4, c5]:
                    cell.border = thin_border
                    cell.font = font_body
                    cell.alignment = align_left
                c1.alignment = align_center
                c4.alignment = align_center
                
                text_len = len(str(comp["answer"]))
                ws.row_dimensions[row_idx].height = max(26, min(140, int(text_len / 40 * 18) + 18))
                row_idx += 1
                
        # 5. 組織卡關點與未來價值展望
        if se.get("reflection"):
            ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=5)
            refl_banner = ws.cell(row_idx, 1, value="💡 組織卡關點反思與未來價值展望")
            refl_banner.font = Font(name="微軟正黑體", size=10, bold=True, color="4E342E")
            refl_banner.fill = fill_sub_header
            refl_banner.alignment = Alignment(horizontal="left", vertical="center", indent=1)
            ws.row_dimensions[row_idx].height = 24
            row_idx += 1
            
            for r_title, r_ans in se["reflection"].items():
                c1 = ws.cell(row_idx, 1, value="卡關與展望")
                c2 = ws.cell(row_idx, 2, value=r_title)
                c3 = ws.cell(row_idx, 3, value=r_ans or "（無填寫）")
                c4 = ws.cell(row_idx, 4, value="")
                c5 = ws.cell(row_idx, 5, value="")
                
                for cell in [c1, c2, c3, c4, c5]:
                    cell.border = thin_border
                    cell.font = font_body
                    cell.alignment = align_left
                c1.alignment = align_center
                c4.alignment = align_center
                
                text_len = len(str(r_ans))
                ws.row_dimensions[row_idx].height = max(26, min(140, int(text_len / 40 * 18) + 18))
                row_idx += 1
                
        row_idx += 2 # gap between chunks
        
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 34
    ws.column_dimensions['C'].width = 62
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 30

# Generate standalone files
for sup_key, members in SUPERVISOR_TEAMS.items():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "部屬自評彙整表"
    create_supervisor_chunk_sheet(ws, sup_key, members, entries)
    fname = f"自評_{sup_key}主管專用_部屬自評彙整表.xlsx" if sup_key != "張希慈_執行長" else "自評_張希慈執行長自評表.xlsx"
    wb.save(fname)
    print(f"Generated {fname}")

# Generate Full Workbook
wb_full = openpyxl.Workbook()
wb_full.remove(wb_full.active)

for sup_key, members in SUPERVISOR_TEAMS.items():
    s_title = f"自評_{sup_key}" if sup_key != "張希慈_執行長" else "自評_執行長"
    ws = wb_full.create_sheet(title=s_title)
    create_supervisor_chunk_sheet(ws, sup_key, members, entries)

# Add "評主管總表"
ws_sup = wb_full.create_sheet(title="評主管總表")
ws_sup.views.sheetView[0].showGridLines = True
ws_sup.cell(1, 1, value="好好星球文化基金會 360 年中成長評估 - 【評主管】回覆總表").font = Font(name="微軟正黑體", size=13, bold=True, color="3E2723")
ws_sup.cell(1, 1).fill = fill_chunk_header
ws_sup.row_dimensions[1].height = 30

sup_headers = [
    "時間戳記", "填答者Email", "受評主管", 
    "Q4.尋求協助", "Q5.具體引導", "Q6.改善幅度", "Q7.跨部門推動",
    "Q8.資源評估", "Q9.失誤回應", "Q10.肯定認可", "Q11.工作成效",
    "Q12.【信任】表達想法", "Q13.【多元】聆聽意見", "Q14.【實驗】嘗試創新",
    "Q15.【實驗】試錯空間", "Q16.【可持續】讚美肯定", "Q17.【可持續】尊重界線",
    "Q18.NPS推薦度", "Q19.滿意度",
    "Q20.願景使命引導(質化)", "Q21.管理提升建議(質化)", "Q22.其他補充評價(質化)", "Q23.肯定感謝詞(星光大賞)"
]
for col_i, h in enumerate(sup_headers, 1):
    c = ws_sup.cell(2, col_i, value=h)
    c.font = font_sub_header
    c.fill = fill_sub_header
    c.alignment = align_header
    c.border = thin_border
ws_sup.row_dimensions[2].height = 26

r_idx = 3
for e in entries:
    if e["relation"] == "主管":
        se = e.get("supervisor_eval", {})
        row_vals = [
            e["timestamp"], e["email"], e["target"],
            se.get("q4_help_easy"), se.get("q5_guidance_freq"), se.get("q6_improve_degree"), se.get("q7_cross_dept"),
            se.get("q8_resource_eval"), se.get("q9_constructive_mistake"), se.get("q10_recognition"), se.get("q11_overall_performance"),
            se.get("q12_trust_express"), se.get("q13_diversity_listen"), se.get("q14_experiment_try"),
            se.get("q15_experiment_psych_safety"), se.get("q16_sustain_praise"), se.get("q17_sustain_boundary"),
            se.get("q18_nps_recommend"), se.get("q19_satisfaction"),
            se.get("q20_vision_mission"), se.get("q21_improvement_advice"), se.get("q22_other_comments"), se.get("q23_starlight_thanks")
        ]
        for col_i, val in enumerate(row_vals, 1):
            c = ws_sup.cell(r_idx, col_i, value=val)
            c.font = font_body
            c.border = thin_border
            c.alignment = align_center if col_i <= 19 else align_left
        ws_sup.row_dimensions[r_idx].height = 40
        r_idx += 1

for c in range(1, len(sup_headers) + 1):
    ws_sup.column_dimensions[get_column_letter(c)].width = 15 if c <= 19 else 32

# Add "評同事總表"
ws_peer = wb_full.create_sheet(title="評同事總表")
ws_peer.views.sheetView[0].showGridLines = True
ws_peer.cell(1, 1, value="好好星球文化基金會 360 年中成長評估 - 【評同事】回覆總表").font = Font(name="微軟正黑體", size=13, bold=True, color="3E2723")
ws_peer.cell(1, 1).fill = fill_chunk_header
ws_peer.row_dimensions[1].height = 30

peer_headers = [
    "時間戳記", "填答者Email", "受評同事",
    "Q24.合作狀況", "Q25.注重細節", "Q26.準時完成", "Q27.靈活調整", "Q28.追蹤承諾", "Q29.說明依據",
    "Q30.【多元】接受不同意見", "Q31.【多元】建設性觀點", "Q32.【實驗】開放調整",
    "Q33.【信任】分享經驗", "Q34.【可持續】讚美同事", "Q35.【可持續】尊重界線",
    "Q36.NPS推薦度",
    "Q37.提升建議(質化)", "Q38.其他評價(質化)", "Q39.肯定感謝詞(星光大賞)"
]
for col_i, h in enumerate(peer_headers, 1):
    c = ws_peer.cell(2, col_i, value=h)
    c.font = font_sub_header
    c.fill = fill_sub_header
    c.alignment = align_header
    c.border = thin_border
ws_peer.row_dimensions[2].height = 26

r_idx = 3
for e in entries:
    if e["relation"] == "同事":
        pe = e.get("peer_eval", {})
        row_vals = [
            e["timestamp"], e["email"], e["target"],
            pe.get("q24_cooperation"), pe.get("q25_detail_oriented"), pe.get("q26_on_time"), pe.get("q27_flexibility"),
            pe.get("q28_follow_up"), pe.get("q29_transparency"),
            pe.get("q30_open_to_opposing"), pe.get("q31_constructive_opinions"), pe.get("q32_growth_mindset"),
            pe.get("q33_share_knowledge"), pe.get("q34_praise_peers"), pe.get("q35_boundary_respect"),
            pe.get("q36_nps_recommend"),
            pe.get("q37_improvement_advice"), pe.get("q38_other_comments"), pe.get("q39_starlight_thanks")
        ]
        for col_i, val in enumerate(row_vals, 1):
            c = ws_peer.cell(r_idx, col_i, value=val)
            c.font = font_body
            c.border = thin_border
            c.alignment = align_center if col_i <= 16 else align_left
        ws_peer.row_dimensions[r_idx].height = 40
        r_idx += 1

for c in range(1, len(peer_headers) + 1):
    ws_peer.column_dimensions[get_column_letter(c)].width = 15 if c <= 16 else 32

wb_full.save("好好星球_360年中成長評估_主管分流與完整彙整表.xlsx")
print("Saved 好好星球_360年中成長評估_主管分流與完整彙整表.xlsx with chunk styling!")
